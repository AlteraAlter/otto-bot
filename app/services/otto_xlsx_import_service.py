"""Helpers for importing OTTO XLSX exports and matching account EANs."""

from __future__ import annotations

import asyncio
import re
from datetime import date, datetime
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterator
from urllib.parse import unquote, urlparse

import httpx
from openpyxl import load_workbook
from sqlalchemy import and_, bindparam, delete, func, or_, select, update
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.otto_xlsx_import import (
    OttoXlsxEanMapping,
    OttoXlsxImportRow,
    OttoXlsxNameMapping,
)
from app.models.product_image_cache import ProductImageCache
import app.models.product_variants  # noqa: F401
from app.models.products import Product
from app.schemas.enums import Controller

XLSX_IMPORT_COLUMN_MAP = {
    "Produktreferenz": "product_reference",
    "SKU": "sku",
    "EAN": "ean",
    "MOIN": "moin",
    "Produktkategorie": "product_category",
    "Marke (ID)": "brand_id",
    "Marke": "brand",
    "Preis": "price",
    "UVP": "recommended_retail_price",
    "Marktplatz-Status": "marketplace_status",
    "Aktiv-Status": "active_status",
    "Link zu otto.de": "otto_url",
}
REQUIRED_XLSX_IMPORT_COLUMNS = [
    "Produktreferenz",
    "SKU",
    "EAN",
    "Produktkategorie",
    "Link zu otto.de",
]
ACCOUNT_NAMES = {"jv", "xl"}
INVALID_MAPPING_TEXT_MARKERS = (
    "product removed",
    "nicht mehr",
    "nichtmehr",
    "no longer",
    "not available",
)
NAME_STOPWORDS = {
    "and",
    "der",
    "die",
    "das",
    "den",
    "dem",
    "des",
    "ein",
    "eine",
    "einer",
    "eines",
    "fuer",
    "made",
    "mit",
    "neu",
    "und",
    "von",
    "xlmoebel",
    "jvmoebel",
    "xl",
    "jv",
}


def _not_product_removed(column: Any) -> Any:
    return and_(
        *[
            or_(column.is_(None), ~column.ilike(f"%{marker}%"))
            for marker in INVALID_MAPPING_TEXT_MARKERS
        ]
    )


def valid_import_row_filters() -> tuple[Any, ...]:
    return (
        _not_product_removed(OttoXlsxImportRow.marketplace_status),
        _not_product_removed(OttoXlsxImportRow.active_status),
        _not_product_removed(OttoXlsxImportRow.product_name),
        _not_product_removed(OttoXlsxImportRow.url_product_name),
    )


def has_removed_marker(value: Any) -> bool:
    text = clean_xlsx_text(value)
    return bool(
        text
        and any(marker in text.casefold() for marker in INVALID_MAPPING_TEXT_MARKERS)
    )


def api_product_is_valid_for_mapping(product: dict[str, Any]) -> bool:
    description = product.get("productDescription")
    description_values = description.values() if isinstance(description, dict) else []
    return not any(
        has_removed_marker(value)
        for value in (
            product.get("marketplaceStatus"),
            product.get("activeStatus"),
            product.get("status"),
            product.get("productLine"),
            product.get("title"),
            product.get("name"),
            *description_values,
        )
    )


def clean_xlsx_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def parse_xlsx_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\u00a0", "")
    if not text:
        return None
    normalized = re.sub(r"[^\d,.-]", "", text)
    if "," in normalized and "." in normalized:
        if normalized.rfind(",") > normalized.rfind("."):
            normalized = normalized.replace(".", "").replace(",", ".")
        else:
            normalized = normalized.replace(",", "")
    elif "," in normalized:
        normalized = normalized.replace(",", ".")
    try:
        return float(normalized)
    except ValueError:
        return None


def normalize_match_name(value: Any) -> str | None:
    text = clean_xlsx_text(value)
    if not text:
        return None
    text = unquote(text).casefold()
    text = (
        text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    tokens = [token for token in text.split() if token not in NAME_STOPWORDS]
    return " ".join(tokens) or None


def match_name_tokens(value: Any) -> set[str]:
    normalized = normalize_match_name(value)
    if not normalized:
        return set()
    return {
        token
        for token in normalized.split()
        if len(token) > 2 and not token.isdigit() and token not in NAME_STOPWORDS
    }


def extract_url_product_name(value: Any) -> str | None:
    url = clean_xlsx_text(value)
    if not url:
        return None
    parsed = urlparse(url)
    segments = [segment for segment in parsed.path.split("/") if segment]
    if len(segments) < 2 or segments[0] != "p":
        return None
    slug = segments[1]
    slug = re.sub(r"-S[A-Z0-9]+$", "", slug, flags=re.IGNORECASE)
    slug = slug.replace("-", " ")
    return clean_xlsx_text(unquote(slug))


def product_match_name(product: dict[str, Any]) -> str | None:
    description = product.get("productDescription")
    if isinstance(description, dict):
        for key in ("productLine", "title", "name"):
            if text := clean_xlsx_text(description.get(key)):
                return text
    for key in ("productLine", "title", "name", "Artikelbeschreibung"):
        if text := clean_xlsx_text(product.get(key)):
            return text
    return None


def row_match_name(row: Any) -> str | None:
    return clean_xlsx_text(
        getattr(row, "product_name", None) or getattr(row, "url_product_name", None)
    )


def product_match_category(product: dict[str, Any]) -> str | None:
    description = product.get("productDescription")
    if isinstance(description, dict):
        if text := clean_xlsx_text(description.get("category")):
            return text
    return clean_xlsx_text(product.get("product_category") or product.get("category"))


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, list):
        return [json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    return value


def normalize_import_row(
    *,
    account: str,
    source_file: str,
    source_row: int,
    row: dict[str, Any],
) -> dict[str, Any]:
    otto_url = clean_xlsx_text(row.get("Link zu otto.de"))
    url_product_name = extract_url_product_name(otto_url)
    normalized_name = normalize_match_name(url_product_name)
    normalized: dict[str, Any] = {
        "account": account.lower(),
        "source_file": source_file,
        "source_row": source_row,
        "raw_payload": json_safe(row),
        "url_product_name": url_product_name,
        "product_name": url_product_name,
        "name_source": "url_slug" if url_product_name else None,
        "normalized_name": normalized_name,
    }
    for header, column in XLSX_IMPORT_COLUMN_MAP.items():
        value = row.get(header)
        normalized[column] = (
            parse_xlsx_float(value)
            if column in {"price", "recommended_retail_price"}
            else clean_xlsx_text(value)
        )
    return normalized


def iter_xlsx_import_rows(path: Path, *, account: str) -> Iterator[dict[str, Any]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        header_row_index: int | None = None
        headers: list[Any] = []
        for index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=5, values_only=True),
            start=1,
        ):
            row_values = list(row)
            if all(column in row_values for column in REQUIRED_XLSX_IMPORT_COLUMNS):
                header_row_index = index
                headers = row_values
                break
        if header_row_index is None:
            raise ValueError(f"Could not find OTTO XLSX header row in {path}")

        for source_row, row in enumerate(
            worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            row_dict = {
                str(header): value
                for header, value in zip(headers, row)
                if header is not None
            }
            if not any(clean_xlsx_text(value) for value in row_dict.values()):
                continue
            yield normalize_import_row(
                account=account,
                source_file=path.as_posix(),
                source_row=source_row,
                row=row_dict,
            )
    finally:
        workbook.close()


async def upsert_xlsx_import_rows(
    session: AsyncSession,
    rows: list[dict[str, Any]],
    *,
    batch_size: int = 1000,
) -> int:
    if not rows:
        return 0

    upserted = 0
    safe_batch_size = min(batch_size, 1000)
    table = OttoXlsxImportRow.__table__
    update_columns = [
        column.name
        for column in table.columns
        if column.name
        not in {"id", "account", "source_file", "source_row", "imported_at"}
    ]
    for start in range(0, len(rows), safe_batch_size):
        chunk = rows[start : start + safe_batch_size]
        stmt = insert(table).values(chunk)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_otto_xlsx_import_account_file_row",
            set_={column: getattr(stmt.excluded, column) for column in update_columns},
        )
        await session.execute(stmt)
        await session.commit()
        upserted += len(chunk)
    return upserted


async def replace_account_rows(session: AsyncSession, account: str) -> int:
    result = await session.execute(
        delete(OttoXlsxImportRow).where(OttoXlsxImportRow.account == account.lower())
    )
    await session.commit()
    return int(result.rowcount or 0)


def response_items(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []
    for key in ("productVariations", "items", "products", "data", "content"):
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def response_total(payload: Any) -> int | None:
    if not isinstance(payload, dict):
        return None
    for key in ("total", "totalElements", "totalCount", "count"):
        value = payload.get(key)
        if isinstance(value, int):
            return value
        if isinstance(value, str) and value.isdigit():
            return int(value)
    return None


def response_has_next(payload: Any) -> bool | None:
    if not isinstance(payload, dict):
        return None
    links = payload.get("links")
    if not isinstance(links, list):
        return None
    return any(
        isinstance(link, dict) and str(link.get("rel") or "").casefold() == "next"
        for link in links
    )


def identifiers_from_api_product(product: dict[str, Any]) -> list[str]:
    identifiers = [
        product.get("productReference"),
        product.get("product_reference"),
        product.get("reference"),
        product.get("sku"),
        product.get("ean"),
        product.get("moin"),
    ]
    cleaned: list[str] = []
    seen: set[str] = set()
    for value in identifiers:
        text = clean_xlsx_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        cleaned.append(text)
    return cleaned


def product_name_from_api_product(product: dict[str, Any]) -> str | None:
    description = product.get("productDescription")
    if isinstance(description, dict):
        for key in ("productLine", "title", "name"):
            if text := clean_xlsx_text(description.get(key)):
                return text
    for key in ("productLine", "title", "name"):
        if text := clean_xlsx_text(product.get(key)):
            return text
    return None


async def update_import_names_for_product_references(
    session: AsyncSession,
    *,
    account: str,
    names_by_reference: dict[str, str],
    only_missing: bool = True,
) -> int:
    rows = [
        {
            "b_identifier": product_reference,
            "b_product_name": product_name,
            "b_normalized_name": normalize_match_name(product_name),
        }
        for product_reference, product_name in names_by_reference.items()
        if product_reference and product_name
    ]
    if not rows:
        return 0

    table = OttoXlsxImportRow.__table__
    stmt = (
        update(table)
        .where(
            table.c.account == account.lower(),
            or_(
                table.c.product_reference == bindparam("b_identifier"),
                table.c.sku == bindparam("b_identifier"),
                table.c.ean == bindparam("b_identifier"),
                table.c.moin == bindparam("b_identifier"),
            ),
        )
        .values(
            product_name=bindparam("b_product_name"),
            name_source="otto_api_product_reference",
            normalized_name=bindparam("b_normalized_name"),
        )
    )
    if only_missing:
        stmt = stmt.where(table.c.product_name.is_(None))

    result = await session.execute(stmt, rows)
    await session.commit()
    return int(result.rowcount or 0)


async def enrich_import_names_from_otto_api(
    session: AsyncSession,
    *,
    client: Any,
    account: str,
    page_size: int = 1000,
    start_page: int = 0,
    limit_pages: int | None = None,
    only_missing: bool = True,
    max_retries: int = 8,
    retry_base_sleep_seconds: float = 8.0,
) -> dict[str, Any]:
    account = account.lower()
    page = max(0, start_page)
    pages_fetched = 0
    products_seen = 0
    products_with_names = 0
    updated_rows = 0
    total: int | None = None

    while True:
        if limit_pages is not None and pages_fetched >= limit_pages:
            break

        for attempt in range(max_retries + 1):
            try:
                payload = await client.get_products_raw(
                    {"page": page, "limit": page_size},
                    controller=account,
                )
                break
            except httpx.HTTPStatusError as exc:
                status_code = exc.response.status_code
                if (
                    status_code not in {429, 500, 502, 503, 504}
                    or attempt >= max_retries
                ):
                    raise
                retry_after = exc.response.headers.get("retry-after")
                if retry_after and retry_after.isdigit():
                    sleep_seconds = float(retry_after)
                else:
                    sleep_seconds = retry_base_sleep_seconds * (attempt + 1)
                await asyncio.sleep(sleep_seconds)
        if total is None:
            total = response_total(payload)

        products = response_items(payload)
        if not products:
            break

        names_by_reference: dict[str, str] = {}
        for product in products:
            products_seen += 1
            product_name = product_name_from_api_product(product)
            identifiers = identifiers_from_api_product(product)
            if identifiers and product_name:
                for identifier in identifiers:
                    names_by_reference[identifier] = product_name
                products_with_names += 1

        updated_rows += await update_import_names_for_product_references(
            session,
            account=account,
            names_by_reference=names_by_reference,
            only_missing=only_missing,
        )

        pages_fetched += 1
        has_next = response_has_next(payload)
        if has_next is False:
            break
        if total is not None and products_seen >= total:
            break
        if has_next is None and len(products) < page_size:
            break
        page += 1

    return {
        "account": account,
        "start_page": start_page,
        "pages_fetched": pages_fetched,
        "products_seen": products_seen,
        "products_with_names": products_with_names,
        "updated_rows": updated_rows,
        "total_from_otto": total,
        "only_missing": only_missing,
    }


def _mapping_payload(
    *,
    source_account: str,
    target_account: str,
    source: Any,
    target: dict[str, Any] | None,
    status: str,
    strategy: str,
    score: float | None,
    candidate_count: int,
    match_reason: str | None = None,
) -> dict[str, Any]:
    return {
        "source_account": source_account,
        "target_account": target_account,
        "source_import_row_id": source.id,
        "target_import_row_id": target.get("id") if target else None,
        "source_ean": source.ean,
        "target_ean": target.get("ean") if target and status == "matched" else None,
        "source_product_category": source.product_category,
        "target_product_category": target.get("product_category") if target else None,
        "source_name": row_match_name(source),
        "target_name": target.get("product_name") if target else None,
        "source_normalized_name": source.normalized_name,
        "target_normalized_name": target.get("normalized_name") if target else None,
        "score": score,
        "status": status,
        "strategy": strategy,
        "candidate_count": candidate_count,
        "match_reason": match_reason,
    }


def _name_mapping_payload(
    *,
    source_account: str,
    target_account: str,
    source: Any,
    target: dict[str, Any],
    strategy: str,
    score: float,
    candidate_count: int,
    source_row_count: int,
    target_row_count: int,
) -> dict[str, Any]:
    return {
        "source_account": source_account,
        "target_account": target_account,
        "source_import_row_id": source.id,
        "target_import_row_id": target.get("id"),
        "source_ean": clean_xlsx_text(source.ean),
        "target_ean": target.get("ean"),
        "source_product_category": clean_xlsx_text(source.product_category),
        "target_product_category": target.get("product_category"),
        "source_name": row_match_name(source),
        "target_name": target.get("product_name"),
        "source_normalized_name": clean_xlsx_text(source.normalized_name),
        "target_normalized_name": target.get("normalized_name"),
        "source_row_count": source_row_count,
        "target_row_count": target_row_count,
        "score": score,
        "strategy": strategy,
        "candidate_count": candidate_count,
    }


async def upsert_ean_mapping_rows(
    session: AsyncSession,
    rows: list[dict[str, Any]],
    *,
    batch_size: int = 1000,
) -> int:
    if not rows:
        return 0

    upserted = 0
    safe_batch_size = min(batch_size, 1000)
    table = OttoXlsxEanMapping.__table__
    update_columns = [
        column.name
        for column in table.columns
        if column.name
        not in {
            "id",
            "source_account",
            "target_account",
            "source_ean",
            "generated_at",
        }
    ]
    for start in range(0, len(rows), safe_batch_size):
        chunk = rows[start : start + safe_batch_size]
        stmt = insert(table).values(chunk)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_otto_xlsx_ean_mapping_source_target_ean",
            set_={column: getattr(stmt.excluded, column) for column in update_columns},
        )
        await session.execute(stmt)
        await session.commit()
        upserted += len(chunk)
    return upserted


async def upsert_name_mapping_rows(
    session: AsyncSession,
    rows: list[dict[str, Any]],
    *,
    batch_size: int = 1000,
) -> int:
    if not rows:
        return 0

    upserted = 0
    safe_batch_size = min(batch_size, 1000)
    table = OttoXlsxNameMapping.__table__
    update_columns = [
        column.name
        for column in table.columns
        if column.name
        not in {
            "id",
            "source_account",
            "target_account",
            "source_normalized_name",
            "generated_at",
        }
    ]
    for start in range(0, len(rows), safe_batch_size):
        chunk = rows[start : start + safe_batch_size]
        stmt = insert(table).values(chunk)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_otto_xlsx_name_mapping_source_target_name",
            set_={column: getattr(stmt.excluded, column) for column in update_columns},
        )
        await session.execute(stmt)
        await session.commit()
        upserted += len(chunk)
    return upserted


async def rebuild_ean_mappings_by_name(
    session: AsyncSession,
    *,
    source_account: str = "jv",
    target_account: str = "xl",
    minimum_score: float = 0.62,
    batch_size: int = 1000,
) -> dict[str, Any]:
    source_account = source_account.lower()
    target_account = target_account.lower()
    await session.execute(
        delete(OttoXlsxEanMapping).where(
            OttoXlsxEanMapping.source_account == source_account,
            OttoXlsxEanMapping.target_account == target_account,
        )
    )
    await session.commit()

    target_rows = (
        await session.execute(
            select(
                OttoXlsxImportRow.id,
                OttoXlsxImportRow.ean,
                OttoXlsxImportRow.product_category,
                OttoXlsxImportRow.product_name,
                OttoXlsxImportRow.url_product_name,
                OttoXlsxImportRow.normalized_name,
            ).where(
                OttoXlsxImportRow.account == target_account,
                OttoXlsxImportRow.ean.is_not(None),
                OttoXlsxImportRow.normalized_name.is_not(None),
                *valid_import_row_filters(),
            )
        )
    ).all()
    targets: list[dict[str, Any]] = []
    exact_by_category_name: dict[tuple[str | None, str], list[dict[str, Any]]] = (
        defaultdict(list)
    )
    exact_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    token_index: dict[str, list[int]] = defaultdict(list)
    for row in target_rows:
        tokens = match_name_tokens(row.normalized_name)
        if not tokens:
            continue
        target = {
            "id": row.id,
            "ean": clean_xlsx_text(row.ean),
            "product_category": clean_xlsx_text(row.product_category),
            "url_product_name": clean_xlsx_text(row.url_product_name),
            "product_name": row_match_name(row),
            "normalized_name": clean_xlsx_text(row.normalized_name),
            "tokens": tokens,
        }
        targets.append(target)
        target_index = len(targets) - 1
        if target["normalized_name"]:
            exact_by_name[target["normalized_name"]].append(target)
            exact_by_category_name[
                (target["product_category"], target["normalized_name"])
            ].append(target)
        for token in tokens:
            token_index[token].append(target_index)

    common_tokens = {
        token for token, indexes in token_index.items() if len(indexes) > 2500
    }

    source_rows = (
        await session.execute(
            select(
                OttoXlsxImportRow.id,
                OttoXlsxImportRow.ean,
                OttoXlsxImportRow.product_category,
                OttoXlsxImportRow.product_name,
                OttoXlsxImportRow.url_product_name,
                OttoXlsxImportRow.normalized_name,
            )
            .where(
                OttoXlsxImportRow.account == source_account,
                OttoXlsxImportRow.ean.is_not(None),
                *valid_import_row_filters(),
            )
            .order_by(OttoXlsxImportRow.id.asc())
        )
    ).all()

    seen_source_eans: set[str] = set()
    status_counts: dict[str, int] = defaultdict(int)
    batch: list[dict[str, Any]] = []
    upserted = 0

    for source in source_rows:
        source_ean = clean_xlsx_text(source.ean)
        if not source_ean or source_ean in seen_source_eans:
            continue
        seen_source_eans.add(source_ean)

        source_name = clean_xlsx_text(source.normalized_name)
        if not source_name:
            payload = _mapping_payload(
                source_account=source_account,
                target_account=target_account,
                source=source,
                target=None,
                status="no_source_name",
                strategy="xlsx_name",
                score=None,
                candidate_count=0,
                match_reason="JV XLSX row has no product name in OTTO URL slug.",
            )
        else:
            exact_candidates = exact_by_category_name.get(
                (clean_xlsx_text(source.product_category), source_name),
                [],
            )
            strategy = "exact_name_category"
            if not exact_candidates:
                exact_candidates = exact_by_name.get(source_name, [])
                strategy = "exact_name"

            if exact_candidates:
                status = "matched" if len(exact_candidates) == 1 else "ambiguous"
                payload = _mapping_payload(
                    source_account=source_account,
                    target_account=target_account,
                    source=source,
                    target=exact_candidates[0],
                    status=status,
                    strategy=strategy,
                    score=1.0,
                    candidate_count=len(exact_candidates),
                    match_reason=(
                        None
                        if status == "matched"
                        else "Multiple XL rows have the same normalized name."
                    ),
                )
            else:
                source_tokens = match_name_tokens(source_name)
                candidate_indexes: set[int] = set()
                for token in source_tokens - common_tokens:
                    candidate_indexes.update(token_index.get(token, []))

                best: dict[str, Any] | None = None
                best_score = 0.0
                best_count = 0
                for candidate_index in candidate_indexes:
                    candidate = targets[candidate_index]
                    overlap = source_tokens & candidate["tokens"]
                    if not overlap:
                        continue
                    score = len(overlap) / max(
                        len(source_tokens),
                        len(candidate["tokens"]),
                    )
                    if clean_xlsx_text(source.product_category) == candidate.get(
                        "product_category"
                    ):
                        score += 0.08
                    score = min(score, 1.0)
                    if score > best_score:
                        best = candidate
                        best_score = score
                        best_count = 1
                    elif score == best_score:
                        best_count += 1

                if best is None:
                    payload = _mapping_payload(
                        source_account=source_account,
                        target_account=target_account,
                        source=source,
                        target=None,
                        status="no_target_candidate",
                        strategy="fuzzy_name",
                        score=None,
                        candidate_count=0,
                        match_reason="No XL name shares usable tokens with this JV name.",
                    )
                elif best_score < minimum_score:
                    payload = _mapping_payload(
                        source_account=source_account,
                        target_account=target_account,
                        source=source,
                        target=best,
                        status="low_score",
                        strategy="fuzzy_name",
                        score=best_score,
                        candidate_count=len(candidate_indexes),
                        match_reason="Best XL candidate is below minimum score.",
                    )
                else:
                    status = "matched" if best_count == 1 else "ambiguous"
                    payload = _mapping_payload(
                        source_account=source_account,
                        target_account=target_account,
                        source=source,
                        target=best,
                        status=status,
                        strategy="fuzzy_name",
                        score=best_score,
                        candidate_count=len(candidate_indexes),
                        match_reason=(
                            None
                            if status == "matched"
                            else "Multiple XL rows have the same best fuzzy score."
                        ),
                    )

        status_counts[payload["status"]] += 1
        batch.append(payload)
        if len(batch) >= batch_size:
            upserted += await upsert_ean_mapping_rows(
                session,
                batch,
                batch_size=batch_size,
            )
            batch = []

    if batch:
        upserted += await upsert_ean_mapping_rows(
            session,
            batch,
            batch_size=batch_size,
        )

    return {
        "source_account": source_account,
        "target_account": target_account,
        "source_rows": len(source_rows),
        "unique_source_eans": len(seen_source_eans),
        "target_candidates": len(targets),
        "inserted": upserted,
        "status_counts": dict(sorted(status_counts.items())),
        "minimum_score": minimum_score,
    }


async def rebuild_name_mappings_by_name(
    session: AsyncSession,
    *,
    source_account: str = "jv",
    target_account: str = "xl",
    minimum_score: float = 0.62,
    batch_size: int = 1000,
) -> dict[str, Any]:
    source_account = source_account.lower()
    target_account = target_account.lower()
    await session.execute(
        delete(OttoXlsxNameMapping).where(
            OttoXlsxNameMapping.source_account == source_account,
            OttoXlsxNameMapping.target_account == target_account,
        )
    )
    await session.commit()

    target_rows = (
        await session.execute(
            select(
                OttoXlsxImportRow.id,
                OttoXlsxImportRow.ean,
                OttoXlsxImportRow.product_category,
                OttoXlsxImportRow.product_name,
                OttoXlsxImportRow.url_product_name,
                OttoXlsxImportRow.normalized_name,
            ).where(
                OttoXlsxImportRow.account == target_account,
                OttoXlsxImportRow.normalized_name.is_not(None),
                *valid_import_row_filters(),
            )
        )
    ).all()
    targets: list[dict[str, Any]] = []
    exact_by_category_name: dict[tuple[str | None, str], list[dict[str, Any]]] = (
        defaultdict(list)
    )
    exact_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    token_index: dict[str, list[int]] = defaultdict(list)
    for row in target_rows:
        tokens = match_name_tokens(row.normalized_name)
        normalized_name = clean_xlsx_text(row.normalized_name)
        target_name = row_match_name(row)
        if not tokens or not normalized_name or not target_name:
            continue
        target = {
            "id": row.id,
            "ean": clean_xlsx_text(row.ean),
            "product_category": clean_xlsx_text(row.product_category),
            "url_product_name": clean_xlsx_text(row.url_product_name),
            "product_name": target_name,
            "normalized_name": normalized_name,
            "tokens": tokens,
        }
        targets.append(target)
        target_index = len(targets) - 1
        exact_by_name[normalized_name].append(target)
        exact_by_category_name[(target["product_category"], normalized_name)].append(
            target
        )
        for token in tokens:
            token_index[token].append(target_index)

    common_tokens = {
        token for token, indexes in token_index.items() if len(indexes) > 2500
    }

    source_rows = (
        await session.execute(
            select(
                OttoXlsxImportRow.id,
                OttoXlsxImportRow.ean,
                OttoXlsxImportRow.product_category,
                OttoXlsxImportRow.product_name,
                OttoXlsxImportRow.url_product_name,
                OttoXlsxImportRow.normalized_name,
            )
            .where(
                OttoXlsxImportRow.account == source_account,
                OttoXlsxImportRow.normalized_name.is_not(None),
                *valid_import_row_filters(),
            )
            .order_by(OttoXlsxImportRow.id.asc())
        )
    ).all()

    source_groups: dict[str, list[Any]] = defaultdict(list)
    for source in source_rows:
        source_name = clean_xlsx_text(source.normalized_name)
        if source_name and row_match_name(source):
            source_groups[source_name].append(source)

    status_counts: dict[str, int] = defaultdict(int)
    batch: list[dict[str, Any]] = []
    upserted = 0

    for source_name, rows in source_groups.items():
        source = rows[0]
        exact_candidates = exact_by_category_name.get(
            (clean_xlsx_text(source.product_category), source_name),
            [],
        )
        strategy = "exact_name_category"
        if not exact_candidates:
            exact_candidates = exact_by_name.get(source_name, [])
            strategy = "exact_name"

        if exact_candidates:
            if len(exact_candidates) != 1:
                status_counts["ambiguous"] += 1
                continue
            target = exact_candidates[0]
            score = 1.0
            candidate_count = 1
        else:
            source_tokens = match_name_tokens(source_name)
            candidate_indexes: set[int] = set()
            for token in source_tokens - common_tokens:
                candidate_indexes.update(token_index.get(token, []))

            best: dict[str, Any] | None = None
            best_score = 0.0
            best_count = 0
            for candidate_index in candidate_indexes:
                candidate = targets[candidate_index]
                overlap = source_tokens & candidate["tokens"]
                if not overlap:
                    continue
                score = len(overlap) / max(
                    len(source_tokens),
                    len(candidate["tokens"]),
                )
                if clean_xlsx_text(source.product_category) == candidate.get(
                    "product_category"
                ):
                    score += 0.08
                score = min(score, 1.0)
                if score > best_score:
                    best = candidate
                    best_score = score
                    best_count = 1
                elif score == best_score:
                    best_count += 1

            if best is None:
                status_counts["no_target_candidate"] += 1
                continue
            if best_score < minimum_score:
                status_counts["low_score"] += 1
                continue
            if best_count != 1:
                status_counts["ambiguous"] += 1
                continue
            target = best
            score = best_score
            strategy = "fuzzy_name"
            candidate_count = len(candidate_indexes)

        target_normalized_name = target.get("normalized_name")
        if not target_normalized_name:
            status_counts["invalid_target_name"] += 1
            continue
        status_counts["matched"] += 1
        batch.append(
            _name_mapping_payload(
                source_account=source_account,
                target_account=target_account,
                source=source,
                target=target,
                strategy=strategy,
                score=score,
                candidate_count=candidate_count,
                source_row_count=len(rows),
                target_row_count=len(exact_by_name.get(target_normalized_name, [])),
            )
        )
        if len(batch) >= batch_size:
            upserted += await upsert_name_mapping_rows(
                session,
                batch,
                batch_size=batch_size,
            )
            batch = []

    if batch:
        upserted += await upsert_name_mapping_rows(
            session,
            batch,
            batch_size=batch_size,
        )

    return {
        "source_account": source_account,
        "target_account": target_account,
        "source_rows": len(source_rows),
        "unique_source_names": len(source_groups),
        "target_candidates": len(targets),
        "inserted": upserted,
        "status_counts": dict(sorted(status_counts.items())),
        "minimum_score": minimum_score,
    }


async def map_eans_from_import_by_name(
    session: AsyncSession,
    *,
    target_controller: Controller,
    products: list[dict[str, Any]],
    minimum_score: float = 0.62,
) -> tuple[dict[str, str], dict[str, Any]]:
    source_items: list[dict[str, Any]] = []
    categories: set[str] = set()
    for product in products:
        if not isinstance(product, dict):
            continue
        if not api_product_is_valid_for_mapping(product):
            continue
        source_ean = clean_xlsx_text(product.get("ean"))
        source_name = product_match_name(product)
        source_tokens = match_name_tokens(source_name)
        if not source_ean or not source_name or len(source_tokens) < 2:
            continue
        category = product_match_category(product)
        if category:
            categories.add(category)
        source_items.append(
            {
                "ean": source_ean,
                "name": source_name,
                "tokens": source_tokens,
                "category": category,
            }
        )
    if not source_items:
        return {}, {"strategy": "xlsx_name", "reason": "no_source_names"}

    stmt = select(
        OttoXlsxImportRow.ean,
        OttoXlsxImportRow.product_category,
        OttoXlsxImportRow.url_product_name,
        OttoXlsxImportRow.normalized_name,
        OttoXlsxImportRow.source_file,
        OttoXlsxImportRow.source_row,
    ).where(
        OttoXlsxImportRow.account == target_controller.value.lower(),
        OttoXlsxImportRow.ean.is_not(None),
        OttoXlsxImportRow.normalized_name.is_not(None),
        *valid_import_row_filters(),
    )

    filtered_stmt = stmt
    if categories:
        filtered_stmt = filtered_stmt.where(
            OttoXlsxImportRow.product_category.in_(categories)
        )

    rows = (await session.execute(filtered_stmt)).all()
    if categories and not rows:
        rows = (await session.execute(stmt)).all()

    candidates: list[dict[str, Any]] = []
    for row in rows:
        name = row.normalized_name or row.url_product_name
        tokens = match_name_tokens(name)
        if not tokens:
            continue
        candidates.append(
            {
                "ean": clean_xlsx_text(row.ean),
                "category": clean_xlsx_text(row.product_category),
                "name": clean_xlsx_text(row.url_product_name),
                "tokens": tokens,
                "source_file": row.source_file,
                "source_row": row.source_row,
            }
        )

    mapped: dict[str, str] = {}
    samples: list[dict[str, Any]] = []
    ambiguous = 0
    for source in source_items:
        best: dict[str, Any] | None = None
        best_score = 0.0
        best_tie = False
        for candidate in candidates:
            overlap = source["tokens"] & candidate["tokens"]
            if not overlap:
                continue
            score = len(overlap) / max(len(source["tokens"]), len(candidate["tokens"]))
            if source.get("category") and source["category"] == candidate.get(
                "category"
            ):
                score += 0.08
            score = min(score, 1.0)
            if score > best_score:
                best = candidate
                best_score = score
                best_tie = False
            elif score == best_score and best is not None:
                best_tie = True
        if best is None or best_score < minimum_score or best_tie:
            if best_tie:
                ambiguous += 1
            continue
        target_ean = clean_xlsx_text(best.get("ean"))
        if not target_ean:
            continue
        mapped[source["ean"]] = target_ean
        if len(samples) < 8:
            samples.append(
                {
                    "source_ean": source["ean"],
                    "target_ean": target_ean,
                    "score": round(best_score, 3),
                    "source_name": source["name"],
                    "target_name": best.get("name"),
                    "target_row": best.get("source_row"),
                }
            )

    return mapped, {
        "strategy": "xlsx_name",
        "target_account": target_controller.value,
        "source_products": len(products),
        "source_with_names": len(source_items),
        "target_candidates": len(candidates),
        "mapped": len(mapped),
        "ambiguous": ambiguous,
        "minimum_score": minimum_score,
        "samples": samples,
    }


async def import_otto_xlsx_rows_to_products(
    session: AsyncSession,
    *,
    account: str | None = None,
    batch_size: int = 1000,
) -> dict[str, Any]:
    """Upsert valid imported OTTO XLSX rows into the local `products` table."""
    stmt = (
        select(
            OttoXlsxImportRow.id,
            OttoXlsxImportRow.account,
            OttoXlsxImportRow.product_reference,
            OttoXlsxImportRow.sku,
            OttoXlsxImportRow.ean,
            OttoXlsxImportRow.moin,
            OttoXlsxImportRow.product_category,
            OttoXlsxImportRow.price,
            OttoXlsxImportRow.recommended_retail_price,
            OttoXlsxImportRow.marketplace_status,
            OttoXlsxImportRow.active_status,
            OttoXlsxImportRow.otto_url,
        )
        .where(
            OttoXlsxImportRow.sku.is_not(None),
            *valid_import_row_filters(),
        )
        .order_by(OttoXlsxImportRow.id.asc())
    )
    if account:
        stmt = stmt.where(OttoXlsxImportRow.account == account.lower())

    source_rows = (await session.execute(stmt)).all()
    rows_by_sku: dict[str, Any] = {}
    duplicate_skus = 0
    for row in source_rows:
        sku = clean_xlsx_text(row.sku)
        if not sku:
            continue
        if sku in rows_by_sku:
            duplicate_skus += 1
        rows_by_sku[sku] = row

    if not rows_by_sku:
        return {
            "account": account,
            "source_rows": len(source_rows),
            "unique_skus": 0,
            "duplicate_skus": duplicate_skus,
            "upserted": 0,
            "with_cached_images": 0,
        }

    table = Product.__table__
    update_columns = [
        "product_reference",
        "ean",
        "moin",
        "product_category",
        "price",
        "recommended_retail_price",
        "marketplace_status",
        "active_status",
        "otto_url",
    ]
    upserted = 0
    with_cached_images = 0
    rows = list(rows_by_sku.values())
    safe_batch_size = min(batch_size, 1000)

    for start in range(0, len(rows), safe_batch_size):
        chunk = rows[start : start + safe_batch_size]
        eans = {ean for row in chunk if (ean := clean_xlsx_text(row.ean))}
        media_by_ean: dict[str, list[str]] = {}
        if eans:
            cached_rows = (
                await session.execute(
                    select(
                        ProductImageCache.ean,
                        ProductImageCache.media_asset_links,
                    ).where(ProductImageCache.ean.in_(eans))
                )
            ).all()
            media_by_ean = {ean: links for ean, links in cached_rows if ean and links}

        payloads: list[dict[str, Any]] = []
        for row in chunk:
            ean = clean_xlsx_text(row.ean)
            media_asset_links = media_by_ean.get(ean or "")
            if media_asset_links:
                with_cached_images += 1
            payloads.append(
                {
                    "product_reference": clean_xlsx_text(row.product_reference),
                    "sku": clean_xlsx_text(row.sku),
                    "ean": ean,
                    "moin": clean_xlsx_text(row.moin),
                    "product_category": clean_xlsx_text(row.product_category),
                    "price": row.price,
                    "recommended_retail_price": row.recommended_retail_price,
                    "marketplace_status": clean_xlsx_text(row.marketplace_status),
                    "active_status": clean_xlsx_text(row.active_status),
                    "otto_url": clean_xlsx_text(row.otto_url),
                    "media_asset_links": media_asset_links,
                }
            )

        insert_stmt = insert(table).values(payloads)
        update_values = {
            column: getattr(insert_stmt.excluded, column) for column in update_columns
        }
        update_values["media_asset_links"] = func.coalesce(
            insert_stmt.excluded.media_asset_links,
            table.c.media_asset_links,
        )
        insert_stmt = insert_stmt.on_conflict_do_update(
            constraint="uq_products_sku",
            set_=update_values,
        )
        await session.execute(insert_stmt)
        await session.commit()
        upserted += len(payloads)

    return {
        "account": account,
        "source_rows": len(source_rows),
        "unique_skus": len(rows_by_sku),
        "duplicate_skus": duplicate_skus,
        "upserted": upserted,
        "with_cached_images": with_cached_images,
    }
