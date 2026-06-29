"""Product endpoints for catalog, creation, deletion, and XLSX import."""

import asyncio
import copy
import json
import logging
import re
import xml.etree.ElementTree as ET
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Optional
from uuid import uuid4

import httpx
from fastapi import (
    APIRouter,
    BackgroundTasks,
    Body,
    Depends,
    File,
    HTTPException,
    Query,
    Request,
    UploadFile,
    WebSocket,
    WebSocketDisconnect,
    status,
)
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from sqlalchemy import asc, desc, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.arq_app import enqueue_job
from app.core.configs import settings
from app.core.user_auth import UserAuth
from app.database import SessionLocal, get_db
from app.dependencies import (
    get_afterbuy_login,
    get_product_service,
    require_role,
)
from app.mapper.normalizer import brand_id_for_controller, build_normalized_product
from app.models.attributes import Attribute
from app.models.categories import Category
from app.models.category_group import CategoryGroup
from app.models.factories import Factories
from app.models.factory_task_states import FactoryTaskState
from app.models.product_import_tasks import ProductImportTask
from app.models.products import Product
from app.models.variation_theme import VariationTheme
from app.repository.user_repository import UserRepository
from app.schemas.enums import Controller, RoleEnum, SortOrderEnum
from app.schemas.ean_pool import (
    EanPoolImportRequest,
    EanPoolImportResponse,
    EanPoolItemResponse,
    EanPoolListResponse,
    EanPoolMarkUsedRequest,
    EanPoolReserveRequest,
    EanPoolStatsResponse,
    EanPoolStatus,
)
from app.schemas.product import (
    Availability,
    CreateProductRequest,
)
from app.schemas.product import (
    Product as ProductPayload,
)
from app.schemas.product import ProductResponse
from app.schemas.product_creation import (
    ProductCreationErrorResponse,
    ProductImportTaskDTO,
    ProductImportTaskListResponse,
)
from app.schemas.product_query import CategoryQuery
from app.schemas.product_response import (
    AvailabilityResponse,
    ProductCreateResponse,
)
from app.schemas.product_tasks import (
    ProductFactoryCreateRequestDTO,
    ProductFactoryCreateResponseDTO,
)
from app.services.afterbuy_service import AfterbuyService
from app.services.ean_pool_service import EanPoolService, ean_pool_item_to_dict
from app.services.factory_task_state_service import FactoryTaskStateService
from app.services.product_service import ProductService
from app.services.product_variant_service import (
    ProductVariantService,
    find_identifier_conflicts,
)
from app.services.product_variation_logic import (
    active_variant_items,
    expand_products_with_variants,
    validate_variant_export_identifiers,
)
from app.services.translation_service import TranslationService, normalize_translation_text
from app.services.variant_image_service import generate_variant_image_from_snapshot

router = APIRouter(
    prefix="/v1/products",
    tags=["Products"],
)
otto_v5_router = APIRouter(
    prefix="/v5/products",
    tags=["OTTO Products v5"],
)

XLSX_COLUMN_MAP = {
    "Produktreferenz": "product_reference",
    "SKU": "sku",
    "EAN": "ean",
    "MOIN": "moin",
    "Produktkategorie": "product_category",
    "Lieferzeit": "delivery_time",
    "Preis": "price",
    "UVP": "recommended_retail_price",
    "Sale-Preis": "sale_price",
    "Sale-Start": "sale_start",
    "Sale-Ende": "sale_end",
    "Marktplatz-Status": "marketplace_status",
    "Fehler": "error_message",
    "Aktiv-Status": "active_status",
    "Link zu otto.de": "otto_url",
    "Datum der letzten Änderung": "last_changed_at",
}
REQUIRED_XLSX_COLUMNS = list(XLSX_COLUMN_MAP.keys())
MAX_TASK_ERROR_LENGTH = 280
MAX_OTTO_PRODUCT_LINE_LENGTH = 70
CREATE_PRODUCT_EXCLUDED_FIELDS = {
    "aiCategory",
    "aiCategoryGroup",
    "shippingProfileID",
    "shippingProfileId",
    "shipping_profile_id",
    "quantity",
    "processingTime",
}
ATTRIBUTES_LIST_PATH = Path(__file__).resolve().parents[3] / "attributes_list.txt"
MAPPER_LOG_PATH = (
    Path(__file__).resolve().parents[3] / "logs" / "product_mapper_flow.log"
)
MAPPER_LOGGER = logging.getLogger("product_mapper_flow")
if not MAPPER_LOGGER.handlers:
    MAPPER_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        _handler = logging.FileHandler(MAPPER_LOG_PATH, encoding="utf-8")
    except OSError:
        _handler = logging.NullHandler()
    _handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    MAPPER_LOGGER.setLevel(logging.INFO)
    MAPPER_LOGGER.addHandler(_handler)
    MAPPER_LOGGER.propagate = False

PREPARED_UPLOAD_LOG_PATH = (
    Path(__file__).resolve().parents[3] / "logs" / "prepared_upload_payloads.log"
)
PREPARED_UPLOAD_LOGGER = logging.getLogger("prepared_upload_payloads")
if not PREPARED_UPLOAD_LOGGER.handlers:
    PREPARED_UPLOAD_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    try:
        _prepared_handler = logging.FileHandler(
            PREPARED_UPLOAD_LOG_PATH, encoding="utf-8"
        )
    except OSError:
        _prepared_handler = logging.NullHandler()
    _prepared_handler.setFormatter(
        logging.Formatter("%(asctime)s %(levelname)s %(message)s")
    )
    PREPARED_UPLOAD_LOGGER.setLevel(logging.INFO)
    PREPARED_UPLOAD_LOGGER.addHandler(_prepared_handler)
    PREPARED_UPLOAD_LOGGER.propagate = False

FACTORY_PREPARE_TASKS: dict[str, dict[str, Any]] = {}
FACTORY_HEARTBEAT_INTERVAL_SEC = 5
FACTORY_STUCK_THRESHOLD_SEC = 300
FACTORY_QUEUED_STEPS = {
    "prepare_queued",
    "ai_enrichment_queued",
    "otto_create_queued",
}
FACTORY_FINAL_STEPS = {
    "otto_create_done",
    "availability_done",
    "otto_create_failed",
    "availability_failed",
}
FACTORY_FETCH_TIMEOUT_SEC = 120
FACTORY_MAP_TIMEOUT_SEC = 1800
FACTORY_NORMALIZE_TIMEOUT_SEC = 20
FACTORY_PRODUCT_CONCURRENCY = settings.factory_product_concurrency
FACTORY_AI_ENRICH_ITEM_TIMEOUT_SEC = settings.factory_ai_enrich_item_timeout_seconds
MAX_OTTO_MEDIA_ASSETS = 19
MEDIA_URL_RE = re.compile(r"https?://[^\s,;|\"'<>]+")
FACTORY_OTTO_CREATE_BATCH_SIZE = 100
FACTORY_OTTO_UPDATE_TASK_MAX_POLLS = 60
FACTORY_OTTO_UPDATE_TASK_FALLBACK_SLEEP_SEC = 5
FACTORY_AVAILABILITY_AFTER_CREATE_DELAY_SEC = 8
FACTORY_TASK_STATE_SERVICE = FactoryTaskStateService()
PRODUCT_DEACTIVATE_CONCURRENCY = 10


def _reset_log_file(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        path.write_text("", encoding="utf-8")
    except OSError:
        return


def _flush_logger(logger: logging.Logger) -> None:
    for handler in logger.handlers:
        try:
            handler.flush()
        except Exception:
            continue


async def _save_factory_task_state(
    process_id: str,
    task: dict[str, Any],
    *,
    created_by_user_id: int | None = None,
) -> dict[str, Any]:
    FACTORY_PREPARE_TASKS[process_id] = task
    return await FACTORY_TASK_STATE_SERVICE.save_task(
        process_id,
        task,
        created_by_user_id=created_by_user_id,
    )


async def _get_factory_task_state(process_id: str) -> dict[str, Any] | None:
    persisted = await FACTORY_TASK_STATE_SERVICE.get_task(process_id)
    if persisted is not None:
        FACTORY_PREPARE_TASKS[process_id] = persisted
        return persisted

    cached = FACTORY_PREPARE_TASKS.get(process_id)
    if cached is not None:
        return cached
    return None


async def _get_owned_factory_task_state(
    process_id: str,
    user_id: int,
) -> dict[str, Any] | None:
    task = await _get_factory_task_state(process_id)
    if task is None:
        return None

    owner = task.get("created_by_user_id")
    if owner is None:
        async with SessionLocal() as session:
            record = await session.get(FactoryTaskState, process_id)
            owner = record.created_by_user_id if record is not None else None

    if owner is None:
        task["created_by_user_id"] = user_id
        return await _save_factory_task_state(
            process_id,
            task,
            created_by_user_id=user_id,
        )

    try:
        owner_id = int(owner)
    except (TypeError, ValueError):
        return None

    if owner_id != user_id:
        return None
    return task


async def _get_websocket_user(websocket: WebSocket):
    auth_header = websocket.headers.get("authorization", "")
    token = ""
    if auth_header.lower().startswith("bearer "):
        token = auth_header.split(" ", 1)[1].strip()
    if not token:
        token = websocket.cookies.get("otto_access_token", "")
    if not token:
        return None

    async with SessionLocal() as session:
        auth = UserAuth(
            user_repository=UserRepository(db=session),
            secret_key=settings.jwt_secret_key,
            algorithm=settings.jwt_algorithm,
            access_token_expire_minutes=settings.jwt_access_token_expire_minutes,
        )
        try:
            user = await auth.get_current_user(token)
        except HTTPException:
            return None
    if user.role != RoleEnum.SEO:
        return None
    return user


def _drop_stale_heartbeat_issues(task: dict[str, Any]) -> None:
    issues = task.get("issues")
    if not isinstance(issues, list):
        return
    task["issues"] = [
        issue
        for issue in issues
        if "без heartbeat" not in str(issue)
        and "Процесс был остановлен или потерян" not in str(issue)
    ]


def _has_factory_task_products(task: dict[str, Any]) -> bool:
    products = task.get("products")
    return isinstance(products, list) and bool(products)


def _is_empty_stale_factory_failure(task: dict[str, Any]) -> bool:
    if task.get("status") != "FAILED":
        return False
    if _has_factory_task_products(task):
        return False
    current_step = str(task.get("current_step") or "")
    return bool(task.get("stuck")) or current_step.startswith(("prepare_", "building_category_"))


async def _mark_factory_task_stale_if_needed(
    process_id: str,
    task: dict[str, Any],
) -> tuple[dict[str, Any], float | None, float | None, bool]:
    heartbeat_at_text = task.get("heartbeat_at")
    step_started_at_text = task.get("current_step_started_at")
    now = datetime.now(UTC)
    heartbeat_lag_sec: float | None = None
    step_elapsed_sec: float | None = None
    is_stuck = False

    if isinstance(heartbeat_at_text, str):
        try:
            heartbeat_lag_sec = (
                now - datetime.fromisoformat(heartbeat_at_text)
            ).total_seconds()
        except ValueError:
            heartbeat_lag_sec = None

    if isinstance(step_started_at_text, str):
        try:
            step_elapsed_sec = (
                now - datetime.fromisoformat(step_started_at_text)
            ).total_seconds()
        except ValueError:
            step_elapsed_sec = None

    current_step = str(task.get("current_step") or "")
    progress_completed = int(task.get("progress_completed") or 0)
    progress_total = int(task.get("progress_total") or 0)
    products = task.get("products")
    has_completed_products = (
        isinstance(products, list)
        and bool(products)
        and progress_total > 0
        and progress_completed >= progress_total
    )
    is_ai_enrichment_step = current_step.startswith("ai_enrichment")
    was_stale_failure = (
        task.get("status") == "FAILED"
        and is_ai_enrichment_step
        and task.get("stuck") is True
        and has_completed_products
    )
    if was_stale_failure:
        recovered_at = now.isoformat()
        task["status"] = "DONE"
        task["current_step"] = "ai_enrichment_done"
        task["current_step_started_at"] = recovered_at
        task["updated_at"] = recovered_at
        task["heartbeat_at"] = recovered_at
        task["enriched_at"] = recovered_at
        task["finished_at"] = recovered_at
        task["stuck"] = False
        task["stuck_message"] = None
        task["progress_percent"] = 100
        _drop_stale_heartbeat_issues(task)
        await _save_factory_task_state(process_id, task)
        MAPPER_LOGGER.warning(
            "step=task_stale_failed_recovered_done process_id=%s",
            process_id,
        )
        return task, heartbeat_lag_sec, step_elapsed_sec, False

    if (
        task.get("status") == "IN_PROGRESS"
        and task.get("current_step") not in FACTORY_QUEUED_STEPS
        and isinstance(heartbeat_lag_sec, float)
        and heartbeat_lag_sec > FACTORY_STUCK_THRESHOLD_SEC
    ):
        if has_completed_products and is_ai_enrichment_step:
            finished_at = now.isoformat()
            task["status"] = "DONE"
            task["current_step"] = "ai_enrichment_done"
            task["current_step_started_at"] = finished_at
            task["updated_at"] = finished_at
            task["heartbeat_at"] = finished_at
            task["enriched_at"] = finished_at
            task["finished_at"] = finished_at
            task["stuck"] = False
            task["stuck_message"] = None
            task["progress_percent"] = 100
            _drop_stale_heartbeat_issues(task)
            await _save_factory_task_state(process_id, task)
            MAPPER_LOGGER.warning(
                "step=task_stale_recovered_done process_id=%s heartbeat_lag_sec=%s",
                process_id,
                int(heartbeat_lag_sec),
            )
            return task, heartbeat_lag_sec, step_elapsed_sec, False

        is_stuck = True
        stale_message = (
            f"Процесс был остановлен или потерян на шаге '{task.get('current_step')}' "
            f"после {int(heartbeat_lag_sec)}s без heartbeat."
        )
        task["status"] = "FAILED"
        task["stuck"] = True
        task["stuck_message"] = stale_message
        task["updated_at"] = now.isoformat()
        task["finished_at"] = now.isoformat()
        issues = task.get("issues")
        task["issues"] = (
            [stale_message, *issues] if isinstance(issues, list) else [stale_message]
        )
        await _save_factory_task_state(process_id, task)
        MAPPER_LOGGER.warning(
            "step=task_stale_marked_failed process_id=%s current_step=%s heartbeat_lag_sec=%s",
            process_id,
            task.get("current_step"),
            int(heartbeat_lag_sec),
        )

    return task, heartbeat_lag_sec, step_elapsed_sec, is_stuck


def _shape_attributes_for_schema(attributes: Any) -> list[dict[str, Any]]:
    if not isinstance(attributes, list):
        return []
    shaped: list[dict[str, Any]] = []
    for item in attributes:
        if not isinstance(item, dict):
            continue
        name = item.get("name")
        if not isinstance(name, str) or not name.strip():
            continue
        if "values" in item and isinstance(item.get("values"), list):
            values = [str(v).strip() for v in item.get("values", []) if str(v).strip()]
        else:
            raw_value = item.get("value")
            if isinstance(raw_value, list):
                values = [str(v).strip() for v in raw_value if str(v).strip()]
            elif raw_value is None:
                values = []
            else:
                text = str(raw_value).strip()
                values = [text] if text else []
        if not values:
            continue
        shaped.append(
            {
                "name": name,
                "values": values,
                "additional": bool(item.get("additional", True)),
            }
        )
    return shaped


def _extract_media_asset_locations(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        locations: list[str] = []
        for item in value:
            locations.extend(_extract_media_asset_locations(item))
        return locations
    if isinstance(value, dict):
        locations: list[str] = []
        for key in ("location", "filename", "url", "src"):
            locations.extend(_extract_media_asset_locations(value.get(key)))
        return locations

    text = str(value).strip()
    if not text:
        return []
    matches = MEDIA_URL_RE.findall(text)
    if matches:
        return [match.rstrip(").]};") for match in matches]
    if text.startswith(("http://", "https://")):
        return [text.rstrip(").]};")]
    return []


def _shape_media_assets_for_schema(assets: Any) -> list[dict[str, Any]]:
    if not isinstance(assets, list):
        return []
    shaped: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in assets:
        asset_type = item.get("type") if isinstance(item, dict) else "IMAGE"
        for location in _extract_media_asset_locations(item):
            if location in seen:
                continue
            seen.add(location)
            shaped.append({"type": str(asset_type or "IMAGE"), "location": location})
            if len(shaped) >= MAX_OTTO_MEDIA_ASSETS:
                return shaped
    return shaped


def _shape_product_media_assets_for_otto(product: dict[str, Any]) -> dict[str, Any]:
    shaped = dict(product)
    shaped["mediaAssets"] = _shape_media_assets_for_schema(shaped.get("mediaAssets"))
    description = shaped.get("productDescription")
    if isinstance(description, dict):
        shaped_description = dict(description)
        product_line = str(shaped_description.get("productLine") or "").strip()
        if len(product_line) > MAX_OTTO_PRODUCT_LINE_LENGTH:
            shaped_description["productLine"] = product_line[
                :MAX_OTTO_PRODUCT_LINE_LENGTH
            ].rstrip()
            MAPPER_LOGGER.info(
                "step=trim_product_line_for_otto sku=%s original_length=%s max_length=%s",
                shaped.get("sku") or "-",
                len(product_line),
                MAX_OTTO_PRODUCT_LINE_LENGTH,
            )
        shaped["productDescription"] = shaped_description
    return shaped


def _pick_text(*values: Any) -> str | None:
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return None


def _extract_specifics_text(xml_data: Any, key: str) -> str | None:
    if not isinstance(xml_data, str) or not xml_data.strip():
        return None

    try:
        root = ET.fromstring(xml_data)
    except ET.ParseError:
        return None

    for item in root.iter():
        children = list(item)
        if len(children) < 2:
            continue
        name = (children[0].text or "").strip()
        value = (children[1].text or "").strip()
        if name == key and value:
            return value
    return None


def _ensure_product_identity(
    *,
    normalized: dict[str, Any],
    source_item: dict[str, Any],
    mapped_item: dict[str, Any] | None,
    index: int,
) -> None:
    source_ean = _pick_text(
        source_item.get("EAN"),
        source_item.get("ean"),
        _extract_specifics_text(source_item.get("CustomItemSpecifics"), "EAN"),
        _extract_specifics_text(source_item.get("CustomItemSpecifics"), "ean"),
    )
    if not source_ean:
        fallback = _pick_text(
            normalized.get("ean"),
            normalized.get("sku"),
            normalized.get("productReference"),
            (mapped_item or {}).get("EAN"),
            (mapped_item or {}).get("ean"),
            f"AUTO-EAN-{index + 1}",
        )
        source_ean = fallback

    normalized["ean"] = source_ean
    normalized["sku"] = source_ean
    normalized["productReference"] = source_ean


def _source_item_ean(source_item: Any) -> str:
    if not isinstance(source_item, dict):
        return ""
    return str(
        _pick_text(
            source_item.get("EAN"),
            source_item.get("ean"),
            _extract_specifics_text(source_item.get("CustomItemSpecifics"), "EAN"),
            _extract_specifics_text(source_item.get("CustomItemSpecifics"), "ean"),
        )
        or ""
    ).strip()


def _save_prepared_payloads_snapshot(
    *,
    payloads: list[ProductPayload],
    controller: Controller,
    factory_id: str,
) -> Path:
    temp_dir = Path(__file__).resolve().parents[3] / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    file_path = (
        temp_dir
        / f"processed_products_{controller.value.lower()}_{factory_id}_{stamp}.json"
    )
    serialized = [item.model_dump(mode="json", exclude_none=True) for item in payloads]
    file_path.write_text(
        json.dumps(serialized, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return file_path


def _save_final_edited_payloads_snapshot(
    *,
    payloads: list[dict[str, Any]],
    controller: Controller,
    factory_id: str,
    process_id: str,
) -> Path:
    temp_dir = Path(__file__).resolve().parents[3] / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    file_path = (
        temp_dir
        / f"processed_products_final_{controller.value.lower()}_{factory_id}_{process_id}_{stamp}.json"
    )
    file_path.write_text(
        json.dumps(payloads, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return file_path


def _save_final_availability_snapshot(
    *,
    payloads: list[dict[str, Any]],
    factory_id: str,
    process_id: str,
) -> Path:
    temp_dir = Path(__file__).resolve().parents[3] / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    file_path = (
        temp_dir
        / f"processed_availability_final_{factory_id}_{process_id}_{stamp}.json"
    )
    file_path.write_text(
        json.dumps(payloads, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return file_path


def _translation_lookup_key(value: Any) -> str:
    return str(value or "").strip().casefold()


def _clean_translation(value: Any) -> str | None:
    cleaned = str(value or "").strip()
    return cleaned or None


def _display_with_ru(value: Any, ru_value: Any) -> str:
    original = str(value or "").strip()
    ru = str(ru_value or "").strip()
    if not original:
        return ru
    if not ru or original.casefold() == ru.casefold():
        return original
    return f"{original} ({ru})"


def _collect_attribute_translation_fallbacks(
    attributes: list[Attribute],
) -> tuple[dict[str, dict[str, str]], dict[tuple[str, str], str]]:
    attribute_fallbacks: dict[str, dict[str, str]] = {}
    value_fallbacks: dict[tuple[str, str], str] = {}
    for attr in attributes:
        attr_key = _translation_lookup_key(attr.name)
        if not attr_key:
            continue
        bucket = attribute_fallbacks.setdefault(attr_key, {})
        attr_name_ru = _clean_translation(attr.name_ru)
        if attr_name_ru and "nameRu" not in bucket:
            bucket["nameRu"] = attr_name_ru
        attr_description_ru = _clean_translation(attr.description_ru)
        if attr_description_ru and "descriptionRu" not in bucket:
            bucket["descriptionRu"] = attr_description_ru

        for item in attr.allowed_values:
            value_key = _translation_lookup_key(item.value)
            value_ru = _clean_translation(item.value_ru)
            if value_key and value_ru and (attr_key, value_key) not in value_fallbacks:
                value_fallbacks[(attr_key, value_key)] = value_ru
    return attribute_fallbacks, value_fallbacks


async def _load_attribute_translation_fallbacks_for_names(
    db: AsyncSession,
    names: list[str],
) -> tuple[dict[str, dict[str, str]], dict[tuple[str, str], str]]:
    lowered_names = sorted({name.strip().lower() for name in names if name.strip()})
    if not lowered_names:
        return {}, {}
    result = await db.execute(
        select(Attribute)
        .options(selectinload(Attribute.allowed_values))
        .where(func.lower(Attribute.name).in_(lowered_names))
    )
    return _collect_attribute_translation_fallbacks(
        list(result.scalars().unique().all())
    )


def _attribute_payload(
    attr: Attribute,
    attribute_fallbacks: dict[str, dict[str, str]],
    value_fallbacks: dict[tuple[str, str], str],
    *,
    is_variation_theme: bool | None = None,
) -> dict[str, Any]:
    attr_key = _translation_lookup_key(attr.name)
    fallback = attribute_fallbacks.get(attr_key, {})
    name_ru = _clean_translation(attr.name_ru) or fallback.get("nameRu")
    description_ru = (
        _clean_translation(attr.description_ru) or fallback.get("descriptionRu")
    )
    allowed_values = sorted(
        {item.value for item in attr.allowed_values if item.value},
        key=str.casefold,
    )
    allowed_values_display = sorted(
        [
            {
                "value": item.value,
                "valueRu": _clean_translation(item.value_ru)
                or value_fallbacks.get(
                    (attr_key, _translation_lookup_key(item.value))
                ),
                "displayValue": item.value,
            }
            for item in attr.allowed_values
            if item.value
        ],
        key=lambda item: str(item["displayValue"]).casefold(),
    )
    payload: dict[str, Any] = {
        "id": attr.id,
        "attributeId": attr.id,
        "attributeKey": str(attr.id),
        "name": attr.name,
        "nameRu": name_ru,
        "displayName": _display_with_ru(attr.name, name_ru),
        "description": attr.description,
        "descriptionRu": description_ru,
        "displayDescription": description_ru or attr.description,
        "type": attr.type,
        "multiValue": attr.multi_value,
        "relevance": attr.relevance,
        "unit": attr.unit,
        "allowedValues": allowed_values,
        "allowedValuesDisplay": allowed_values_display,
    }
    if is_variation_theme is not None:
        payload["isVariationTheme"] = is_variation_theme
    return payload


async def _load_category_group_contexts() -> dict[str, dict[str, Any]]:
    async with SessionLocal() as session:
        result = await session.execute(
            select(CategoryGroup)
            .options(
                selectinload(CategoryGroup.categories),
                selectinload(CategoryGroup.attributes).selectinload(
                    Attribute.allowed_values
                ),
            )
            .order_by(CategoryGroup.name.asc())
        )
        groups = result.scalars().unique().all()
        all_attributes = [attr for group in groups for attr in group.attributes]
        attribute_fallbacks, value_fallbacks = _collect_attribute_translation_fallbacks(
            all_attributes
        )

        contexts: dict[str, dict[str, Any]] = {}
        for group in groups:
            categories = sorted(
                [category.name for category in group.categories if category.name],
                key=str.casefold,
            )
            categories_display = sorted(
                [
                    {
                        "name": category.name,
                        "nameRu": category.name_ru,
                        "displayName": category.name,
                    }
                    for category in group.categories
                    if category.name
                ],
                key=lambda item: str(item["name"]).casefold(),
            )
            attributes = []
            for attr in group.attributes:
                attributes.append(
                    _attribute_payload(attr, attribute_fallbacks, value_fallbacks)
                )
            contexts[group.name] = {
                "categoryGroup": group.name,
                "categoryGroupRu": group.name_ru,
                "displayCategoryGroup": group.name,
                "categories": categories,
                "categoriesDisplay": categories_display,
                "attributes": attributes,
            }

        return contexts


async def _build_factory_prepared_products(
    *,
    payload: ProductFactoryCreateRequestDTO,
    afterbuy: AfterbuyService,
    product_service: ProductService,
    mapping_progress_callback: Any | None = None,
    mapping_total_callback: Any | None = None,
    normalized_item_callback: Any | None = None,
) -> tuple[list[ProductPayload], int, int, list[str], list[dict[str, Any]]]:
    from app.mapper.product_mapper import ProductMapper

    issues: list[str] = []
    _reset_log_file(MAPPER_LOG_PATH)
    _reset_log_file(PREPARED_UPLOAD_LOG_PATH)
    MAPPER_LOGGER.info(
        "step=start_create_from_factory controller=%s factory_id=%s",
        payload.controller.value,
        payload.factory_id,
    )

    async def _run_step_with_timeout(
        *,
        step_name: str,
        timeout_sec: int,
        fn,
    ):
        MAPPER_LOGGER.info("step=%s_start timeout_sec=%s", step_name, timeout_sec)
        try:
            result = await asyncio.wait_for(fn(), timeout=timeout_sec)
            MAPPER_LOGGER.info("step=%s_done", step_name)
            return result
        except asyncio.TimeoutError as exc:
            MAPPER_LOGGER.error(
                "step=%s_timeout timeout_sec=%s", step_name, timeout_sec
            )
            raise RuntimeError(
                f"Step '{step_name}' exceeded timeout ({timeout_sec}s)"
            ) from exc

    source = await _run_step_with_timeout(
        step_name="fetch_afterbuy_products",
        timeout_sec=FACTORY_FETCH_TIMEOUT_SEC,
        fn=lambda: afterbuy.get_products_by_factory_id(
            payload.controller,
            int(payload.factory_id),
        ),
    )
    source_items = [
        item.model_dump(mode="json", exclude_none=True) for item in source.products
    ]

    try:
        await _run_step_with_timeout(
            step_name="fetch_stammartikel_descriptions",
            timeout_sec=FACTORY_FETCH_TIMEOUT_SEC,
            fn=lambda: afterbuy.enrich_items_with_stammartikel_description(
                controller=payload.controller,
                items=source_items,
            ),
        )
    except Exception as exc:
        issues.append(f"stammartikel descriptions skipped: {exc}")
        MAPPER_LOGGER.warning(
            "step=fetch_stammartikel_descriptions_failed error=%s", exc
        )

    if mapping_total_callback is not None:
        await mapping_total_callback(len(source_items))
    MAPPER_LOGGER.info("step=fetch_afterbuy_products count=%s", len(source_items))
    PREPARED_UPLOAD_LOGGER.info(
        "step=fetch_afterbuy_products result=%s",
        json.dumps(
            {"count": len(source_items), "sample": source_items[:2]}, ensure_ascii=False
        ),
    )
    _flush_logger(PREPARED_UPLOAD_LOGGER)

    category_group_contexts = await _run_step_with_timeout(
        step_name="load_local_category_group_contexts",
        timeout_sec=FACTORY_FETCH_TIMEOUT_SEC,
        fn=_load_category_group_contexts,
    )
    MAPPER_LOGGER.info(
        "step=load_local_category_group_contexts groups=%s",
        len(category_group_contexts),
    )

    normalized_results: list[ProductPayload | None] = [None] * len(source_items)
    normalize_semaphore = asyncio.Semaphore(FACTORY_PRODUCT_CONCURRENCY)

    async def _build_and_publish_normalized_one(
        index: int,
        source_item: dict[str, Any],
        mapped_item: dict[str, Any] | None,
    ) -> None:
        try:
            if isinstance(normalized_results[index], ProductPayload):
                return
            ean = source_item.get("EAN") or source_item.get("ean")
            MAPPER_LOGGER.info("step=normalize_start index=%s ean=%s", index, ean)
            async with normalize_semaphore:
                normalized = await asyncio.wait_for(
                    asyncio.to_thread(
                        build_normalized_product,
                        source_item,
                        None,
                        payload.controller,
                    ),
                    timeout=FACTORY_NORMALIZE_TIMEOUT_SEC,
                )
            PREPARED_UPLOAD_LOGGER.info(
                "step=normalize_preview index=%s body=%s",
                index,
                json.dumps(normalized, ensure_ascii=False),
            )
            _flush_logger(PREPARED_UPLOAD_LOGGER)

            product_description = normalized.get("productDescription")
            if isinstance(product_description, dict):
                if isinstance(mapped_item, dict):
                    category_group = mapped_item.get("categoryGroup")
                    if category_group:
                        normalized["aiCategoryGroup"] = category_group
                product_description["category"] = ""
                normalized["aiCategory"] = ""
                product_description["description"] = None
                product_description["bulletPoints"] = []
                product_description["attributes"] = []

            _ensure_product_identity(
                normalized=normalized,
                source_item=source_item,
                mapped_item=mapped_item,
                index=index,
            )
            if isinstance(normalized.get("mediaAssets"), list):
                normalized["mediaAssets"] = _shape_media_assets_for_schema(
                    normalized["mediaAssets"]
                )

            PREPARED_UPLOAD_LOGGER.info(
                "step=final_product_body_candidate index=%s body=%s",
                index,
                json.dumps(normalized, ensure_ascii=False),
            )
            _flush_logger(PREPARED_UPLOAD_LOGGER)

            normalized_model = ProductPayload.model_validate(normalized)
            normalized_results[index] = normalized_model
            if normalized_item_callback is not None:
                await normalized_item_callback(
                    index,
                    normalized_model.model_dump(mode="json", exclude_none=True),
                )
            MAPPER_LOGGER.info(
                "step=normalize_done index=%s ean=%s sku=%s",
                index,
                ean,
                normalized.get("sku"),
            )
        except Exception as exc:
            msg = f"normalize index={index} failed: {exc}"
            issues.append(msg)
            MAPPER_LOGGER.exception("step=normalize_error message=%s", msg)
            PREPARED_UPLOAD_LOGGER.info(
                "step=normalize_error index=%s error=%s",
                index,
                json.dumps(
                    {"message": str(exc), "source_item": source_item},
                    ensure_ascii=False,
                ),
            )
            _flush_logger(PREPARED_UPLOAD_LOGGER)

    async def _publish_mapped_item(index: int, mapped_item: dict[str, Any]) -> None:
        if index >= len(source_items) or not isinstance(source_items[index], dict):
            return
        await _build_and_publish_normalized_one(index, source_items[index], mapped_item)

    mapper = ProductMapper(
        products=source_items,
        controller=payload.controller.value,
        otto_client=product_service.client,
        category_group_contexts=category_group_contexts,
    )
    mapped_result = await _run_step_with_timeout(
        step_name="map_products",
        timeout_sec=FACTORY_MAP_TIMEOUT_SEC,
        fn=lambda: mapper.payload_deploy(
            on_item_finished=mapping_progress_callback,
            on_item_mapped=_publish_mapped_item,
        ),
    )
    mapped_items = (
        mapped_result.get("items", []) if isinstance(mapped_result, dict) else []
    )
    mapped_items_by_index = (
        mapped_result.get("items_by_index", [])
        if isinstance(mapped_result, dict)
        else []
    )
    mapper_issues = (
        mapped_result.get("issues", []) if isinstance(mapped_result, dict) else []
    )
    for item in mapper_issues:
        issues.append(f"mapper index={item.get('index')}: {item.get('message')}")
    MAPPER_LOGGER.info(
        "step=mapper_payload_deploy mapped=%s mapper_issues=%s",
        len(mapped_items),
        len(mapper_issues),
    )
    PREPARED_UPLOAD_LOGGER.info(
        "step=mapper_payload_deploy result=%s",
        json.dumps(
            {
                "mapped_items": len(mapped_items),
                "issues_count": len(mapper_issues),
                "issues": mapper_issues,
                "sample": mapped_items[:2],
            },
            ensure_ascii=False,
        ),
    )
    _flush_logger(PREPARED_UPLOAD_LOGGER)

    async def _normalize_one(index: int, source_item: dict[str, Any]) -> None:
        mapped_item = (
            mapped_items_by_index[index]
            if index < len(mapped_items_by_index)
            and isinstance(mapped_items_by_index[index], dict)
            else None
        )
        if mapped_item is None and normalized_results[index] is None:
            return
        await _build_and_publish_normalized_one(index, source_item, mapped_item)

    await asyncio.gather(
        *[
            _normalize_one(index, source_item)
            for index, source_item in enumerate(source_items)
        ]
    )

    products_payload: list[ProductPayload] = [
        item for item in normalized_results if isinstance(item, ProductPayload)
    ]

    MAPPER_LOGGER.info(
        "step=payload_built payload_items=%s skipped=%s",
        len(products_payload),
        len(source_items) - len(products_payload),
    )
    PREPARED_UPLOAD_LOGGER.info(
        "step=payload_summary result=%s",
        json.dumps(
            {
                "payload_items": len(products_payload),
                "skipped_items": len(source_items) - len(products_payload),
                "issues": issues,
            },
            ensure_ascii=False,
        ),
    )
    _flush_logger(PREPARED_UPLOAD_LOGGER)
    return products_payload, len(source_items), len(mapped_items), issues, source_items


async def _run_factory_prepare_task(
    *,
    process_id: str,
    payload: ProductFactoryCreateRequestDTO,
    afterbuy: AfterbuyService,
    product_service: ProductService,
) -> None:
    started_at = datetime.now(UTC)
    task = await _get_factory_task_state(process_id) or {}
    task.update(
        {
            "status": "IN_PROGRESS",
            "started_at": started_at.isoformat(),
            "updated_at": started_at.isoformat(),
            "heartbeat_at": started_at.isoformat(),
            "heartbeat_count": 0,
            "current_step": "prepare_initializing",
            "current_step_started_at": started_at.isoformat(),
            "stuck": False,
            "stuck_message": None,
            "progress_total": 0,
            "progress_completed": 0,
            "progress_percent": 0,
        }
    )
    await _save_factory_task_state(process_id, task)

    heartbeat_stop = asyncio.Event()

    async def _heartbeat_loop() -> None:
        while not heartbeat_stop.is_set():
            now = datetime.now(UTC)
            current = FACTORY_PREPARE_TASKS.get(process_id)
            if not current:
                await asyncio.sleep(FACTORY_HEARTBEAT_INTERVAL_SEC)
                continue
            current["heartbeat_count"] = int(current.get("heartbeat_count", 0)) + 1
            current["heartbeat_at"] = now.isoformat()
            current["updated_at"] = now.isoformat()
            MAPPER_LOGGER.info(
                "Пульс есть: process_id=%s current_step=%s heartbeat_count=%s",
                process_id,
                current.get("current_step"),
                current["heartbeat_count"],
            )
            await _save_factory_task_state(process_id, current)
            await asyncio.sleep(FACTORY_HEARTBEAT_INTERVAL_SEC)

    heartbeat_task = asyncio.create_task(_heartbeat_loop())

    async def _set_step(step: str) -> None:
        now = datetime.now(UTC)
        current = await _get_factory_task_state(process_id) or {}
        current["current_step"] = step
        current["current_step_started_at"] = now.isoformat()
        current["updated_at"] = now.isoformat()
        await _save_factory_task_state(process_id, current)
        MAPPER_LOGGER.info(
            "step=task_step_change process_id=%s current_step=%s", process_id, step
        )

    progress_lock = asyncio.Lock()

    async def _track_mapping_progress() -> None:
        async with progress_lock:
            current = await _get_factory_task_state(process_id) or {}
            total = int(current.get("progress_total") or 0)
            if total <= 0:
                return
            completed = min(total, int(current.get("progress_completed") or 0) + 1)
            current["progress_completed"] = completed
            current["progress_percent"] = int(round((completed / total) * 100))
            current["updated_at"] = datetime.now(UTC).isoformat()
            await _save_factory_task_state(process_id, current)

    async def _set_mapping_total(total: int) -> None:
        async with progress_lock:
            current = await _get_factory_task_state(process_id) or {}
            current["progress_total"] = max(0, total)
            current["progress_completed"] = 0
            current["progress_percent"] = 0
            current["products"] = []
            current["partial_products_by_index"] = {}
            current["updated_at"] = datetime.now(UTC).isoformat()
            await _save_factory_task_state(process_id, current)

    async def _append_normalized_item(index: int, product: dict[str, Any]) -> None:
        async with progress_lock:
            current = await _get_factory_task_state(process_id) or {}
            by_index_raw = current.get("partial_products_by_index")
            by_index = by_index_raw if isinstance(by_index_raw, dict) else {}
            by_index[str(index)] = product
            current["partial_products_by_index"] = by_index
            current["products"] = [
                by_index[key]
                for key in sorted(by_index, key=lambda value: int(value))
                if isinstance(by_index.get(key), dict)
            ]
            current["payload_items"] = len(current["products"])
            current["updated_at"] = datetime.now(UTC).isoformat()
            await _save_factory_task_state(process_id, current)

    try:
        await _set_step("building_category_preview")
        (
            products_payload,
            source_count,
            mapped_count,
            issues,
            raw_source_items,
        ) = await _build_factory_prepared_products(
            payload=payload,
            afterbuy=afterbuy,
            product_service=product_service,
            mapping_progress_callback=_track_mapping_progress,
            mapping_total_callback=_set_mapping_total,
            normalized_item_callback=_append_normalized_item,
        )
        MAPPER_LOGGER.info(
            "Категории готовы для предпросмотра: process_id=%s source_items=%s mapped_items=%s payload_items=%s",
            process_id,
            source_count,
            mapped_count,
            len(products_payload),
        )
        current = await _get_factory_task_state(process_id) or {}
        current["progress_total"] = source_count
        current["progress_completed"] = min(
            source_count,
            int(current.get("progress_completed") or 0),
        )
        current["progress_percent"] = (
            int(round((current["progress_completed"] / source_count) * 100))
            if source_count > 0
            else 0
        )
        current["updated_at"] = datetime.now(UTC).isoformat()
        await _save_factory_task_state(process_id, current)
        await _set_step("saving_snapshot")
        snapshot_path = _save_prepared_payloads_snapshot(
            payloads=products_payload,
            controller=payload.controller,
            factory_id=payload.factory_id,
        )
        await _save_factory_task_state(
            process_id,
            {
                "status": "DONE",
                "controller": payload.controller.value,
                "factory_id": payload.factory_id,
                "source_items": source_count,
                "mapped_items": mapped_count,
                "payload_items": len(products_payload),
                "issues": issues,
                "source_items_raw": raw_source_items,
                "products": [
                    item.model_dump(mode="json", exclude_none=True)
                    for item in products_payload
                ],
                "partial_products_by_index": {},
                "snapshot_path": snapshot_path.as_posix(),
                "current_step": "category_preview_done",
                "finished_at": datetime.now(UTC).isoformat(),
                "heartbeat_at": datetime.now(UTC).isoformat(),
                "updated_at": datetime.now(UTC).isoformat(),
                "progress_total": source_count,
                "progress_completed": source_count,
                "progress_percent": 100 if source_count > 0 else 0,
            },
        )
    except Exception as exc:
        MAPPER_LOGGER.exception(
            "step=factory_prepare_task_failed process_id=%s error=%s", process_id, exc
        )
        await _save_factory_task_state(
            process_id,
            {
                "status": "FAILED",
                "controller": payload.controller.value,
                "factory_id": payload.factory_id,
                "issues": [str(exc)],
                "products": [],
                "finished_at": datetime.now(UTC).isoformat(),
                "heartbeat_at": datetime.now(UTC).isoformat(),
                "updated_at": datetime.now(UTC).isoformat(),
            },
        )
    finally:
        heartbeat_stop.set()
        try:
            await heartbeat_task
        except Exception:
            pass


def _product_to_dict(product: Product) -> dict[str, Any]:
    """Serialize the local spreadsheet-backed product row."""
    return {
        "id": product.id,
        "productReference": product.product_reference,
        "sku": product.sku,
        "ean": product.ean,
        "moin": product.moin,
        "productCategory": product.product_category,
        "deliveryTime": product.delivery_time,
        "price": product.price,
        "recommendedRetailPrice": product.recommended_retail_price,
        "salePrice": product.sale_price,
        "saleStart": product.sale_start.isoformat() if product.sale_start else None,
        "saleEnd": product.sale_end.isoformat() if product.sale_end else None,
        "marketplaceStatus": product.marketplace_status,
        "errorMessage": product.error_message,
        "activeStatus": product.active_status,
        "ottoUrl": product.otto_url,
        "mediaAssetLinks": product.media_asset_links or [],
        "lastChangedAt": (
            product.last_changed_at.isoformat() if product.last_changed_at else None
        ),
    }


def _summarize_task_error(exc: Exception) -> str:
    """Store a short task error message instead of a full traceback/SQL dump."""
    message = str(exc).strip() or exc.__class__.__name__
    first_line = message.splitlines()[0].strip()
    compact = " ".join(first_line.split())
    if len(compact) <= MAX_TASK_ERROR_LENGTH:
        return compact
    return f"{compact[: MAX_TASK_ERROR_LENGTH - 1].rstrip()}…"


def _task_to_dto(task: ProductImportTask) -> ProductImportTaskDTO:
    return ProductImportTaskDTO(
        id=task.id,
        file_name=task.file_name,
        status=task.status,
        total_rows=task.total_rows,
        processed_rows=task.processed_rows,
        upserted_rows=task.upserted_rows,
        skipped_rows=task.skipped_rows,
        error_message=task.error_message,
        created_at=task.created_at.isoformat(),
        updated_at=task.updated_at.isoformat(),
        started_at=task.started_at.isoformat() if task.started_at else None,
        finished_at=task.finished_at.isoformat() if task.finished_at else None,
    )


def _normalize_ean_lines(raw_items: list[Any]) -> list[str]:
    normalized: list[str] = []
    seen: set[str] = set()
    for item in raw_items:
        text = str(item or "").strip().strip(",").strip('"').strip("'").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        normalized.append(text)
    return normalized


def _extract_process_id(create_result: ProductCreateResponse) -> str | None:
    for link in create_result.links:
        match = re.search(r"update-tasks/([^/?#]+)", link.href)
        if match:
            return match.group(1)
    match = re.search(
        r"\b([0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12})\b",
        create_result.message or "",
        re.IGNORECASE,
    )
    if match:
        return match.group(1)
    return None


def _chunk_list(items: list[Any], size: int) -> list[list[Any]]:
    if size <= 0:
        return [items]
    return [items[index : index + size] for index in range(0, len(items), size)]


def _read_int(value: Any, fallback: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def _compute_update_sleep_seconds(update_result: dict[str, Any]) -> int:
    ping_after_value = update_result.get("pingAfter")
    sleep_seconds = FACTORY_OTTO_UPDATE_TASK_FALLBACK_SLEEP_SEC
    if isinstance(ping_after_value, str):
        try:
            ping_after = datetime.fromisoformat(ping_after_value.replace("Z", "+00:00"))
            now = datetime.now(UTC)
            delta = (ping_after - now).total_seconds()
            sleep_seconds = max(
                1,
                min(
                    FACTORY_OTTO_UPDATE_TASK_FALLBACK_SLEEP_SEC,
                    int(delta) if delta > 0 else 1,
                ),
            )
        except ValueError:
            sleep_seconds = FACTORY_OTTO_UPDATE_TASK_FALLBACK_SLEEP_SEC
    return sleep_seconds


async def _poll_otto_update_task(
    *,
    product_service: ProductService,
    process_id: str,
    controller: Controller,
) -> tuple[dict[str, Any], dict[str, Any] | None, str]:
    update_result: dict[str, Any] = {}
    otto_state = ""

    for _ in range(FACTORY_OTTO_UPDATE_TASK_MAX_POLLS):
        update_result = await product_service.update_tasks(
            process_id, controller=controller
        )
        otto_state = str(update_result.get("state", "")).lower()
        if otto_state in {"done", "failed", "error"}:
            break
        await asyncio.sleep(_compute_update_sleep_seconds(update_result))

    failed_result: dict[str, Any] | None = None
    failed_count = _read_int(update_result.get("failed"))
    if failed_count > 0:
        failed_result = await product_service.failed_tasks(
            process_id, controller=controller
        )

    return update_result, failed_result, otto_state


async def _submit_products_to_otto_in_batches(
    *,
    product_service: ProductService,
    controller: Controller,
    products: list[ProductPayload],
) -> tuple[str | None, str, dict[str, Any], dict[str, Any] | None]:
    batches = _chunk_list(products, FACTORY_OTTO_CREATE_BATCH_SIZE)
    process_ids: list[str] = []
    batch_summaries: list[dict[str, Any]] = []
    aggregated_failed_results: list[Any] = []

    total = 0
    succeeded = 0
    failed = 0
    progress = 0
    final_state = "done"

    for batch_index, batch in enumerate(batches, start=1):
        MAPPER_LOGGER.info(
            "step=submit_final_products_batch_create_start batch=%s batch_size=%s",
            batch_index,
            len(batch),
        )
        create_payload = CreateProductRequest(controller=controller, products=batch)
        create_result = await product_service.create_or_update_products(create_payload)
        otto_process_id = _extract_process_id(create_result)
        create_state = str(create_result.state or "").lower()
        update_result: dict[str, Any] = {}
        failed_result: dict[str, Any] | None = None
        batch_state = create_state

        if otto_process_id:
            process_ids.append(otto_process_id)
            MAPPER_LOGGER.info(
                "step=submit_final_products_batch_poll_start batch=%s otto_process_id=%s",
                batch_index,
                otto_process_id,
            )
            update_result, failed_result, batch_state = await _poll_otto_update_task(
                product_service=product_service,
                process_id=otto_process_id,
                controller=controller,
            )
        else:
            batch_state = "failed"

        batch_total = _read_int(update_result.get("total"), len(batch))
        batch_failed = _read_int(update_result.get("failed"))
        batch_succeeded = _read_int(
            update_result.get("succeeded"),
            max(0, len(batch) - batch_failed) if batch_state == "done" else 0,
        )
        batch_progress = _read_int(
            update_result.get("progress"),
            batch_total if batch_state == "done" else batch_succeeded + batch_failed,
        )

        total += batch_total
        succeeded += batch_succeeded
        failed += batch_failed
        progress += batch_progress

        if batch_state in {"failed", "error"} or batch_failed > 0:
            final_state = "failed"

        if isinstance(failed_result, dict):
            results = failed_result.get("results")
            if isinstance(results, list):
                aggregated_failed_results.extend(results)

        batch_summaries.append(
            {
                "batch": batch_index,
                "batchSize": len(batch),
                "processId": otto_process_id,
                "state": batch_state,
                "total": batch_total,
                "succeeded": batch_succeeded,
                "failed": batch_failed,
                "progress": batch_progress,
            }
        )
        MAPPER_LOGGER.info(
            "step=submit_final_products_batch_done batch=%s process_id=%s failed=%s state=%s",
            batch_index,
            otto_process_id,
            batch_failed,
            batch_state,
        )

    aggregated_update_result = {
        "state": final_state,
        "total": total,
        "progress": min(total, progress),
        "succeeded": succeeded,
        "failed": failed,
        "batchCount": len(batch_summaries),
        "processIds": process_ids,
        "batches": batch_summaries,
    }
    aggregated_failed_result = (
        {"results": aggregated_failed_results} if aggregated_failed_results else None
    )
    aggregated_process_id = ",".join(process_ids) if process_ids else None

    return (
        aggregated_process_id,
        final_state,
        aggregated_update_result,
        aggregated_failed_result,
    )


def _empty_to_none(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    return value


def _as_text(value: Any) -> str | None:
    value = _empty_to_none(value)
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_float(value: Any) -> float | None:
    value = _empty_to_none(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        normalized = value.strip()
        if "," in normalized and "." in normalized:
            normalized = normalized.replace(".", "").replace(",", ".")
        elif "," in normalized:
            normalized = normalized.replace(",", ".")
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def _parse_datetime(value: Any) -> datetime | None:
    value = _empty_to_none(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    if isinstance(value, str):
        for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def _normalize_xlsx_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = {
        "product_reference": _as_text(row.get("Produktreferenz")),
        "sku": _as_text(row.get("SKU")),
        "ean": _as_text(row.get("EAN")),
        "moin": _as_text(row.get("MOIN")),
        "product_category": _as_text(row.get("Produktkategorie")),
        "delivery_time": _as_text(row.get("Lieferzeit")),
        "price": _parse_float(row.get("Preis")),
        "recommended_retail_price": _parse_float(row.get("UVP")),
        "sale_price": _parse_float(row.get("Sale-Preis")),
        "sale_start": _parse_datetime(row.get("Sale-Start")),
        "sale_end": _parse_datetime(row.get("Sale-Ende")),
        "marketplace_status": _as_text(row.get("Marktplatz-Status")),
        "error_message": _as_text(row.get("Fehler")),
        "active_status": _as_text(row.get("Aktiv-Status")),
        "otto_url": _as_text(row.get("Link zu otto.de")),
        "last_changed_at": _parse_datetime(row.get("Datum der letzten Änderung")),
    }
    return normalized


def _read_xlsx_rows(raw: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    worksheet = workbook.active

    header_row_index: int | None = None
    headers: list[Any] = []
    for index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=5, values_only=True),
        start=1,
    ):
        row_values = list(row)
        if all(column in row_values for column in REQUIRED_XLSX_COLUMNS):
            header_row_index = index
            headers = row_values
            break

    if header_row_index is None:
        raise ValueError("Could not find the expected XLSX header row")

    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        row_dict = {
            str(header): value
            for header, value in zip(headers, row)
            if header is not None
        }
        if not any(_empty_to_none(value) is not None for value in row_dict.values()):
            continue
        rows.append(_normalize_xlsx_row(row_dict))

    return rows


def _deduplicate_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    rows_without_identity: list[dict[str, Any]] = []
    rows_by_identity: dict[tuple[str, str | None], dict[str, Any]] = {}

    for row in rows:
        sku = _as_text(row.get("sku"))
        ean = _as_text(row.get("ean"))
        product_reference = _as_text(row.get("product_reference"))

        identity: tuple[str, str | None] | None = None
        if sku:
            identity = ("sku", sku)
        elif ean:
            identity = ("ean", ean)
        elif product_reference:
            identity = ("product_reference", product_reference)

        if identity is None:
            rows_without_identity.append(row)
            continue

        rows_by_identity[identity] = row

    deduplicated_rows = list(rows_by_identity.values()) + rows_without_identity
    skipped_rows = len(rows) - len(deduplicated_rows)
    return deduplicated_rows, skipped_rows


async def _upsert_products_in_batches(
    db: AsyncSession,
    rows: list[dict[str, Any]],
    batch_size: int = 100,
    progress_callback: Any | None = None,
) -> int:
    upserted_rows = 0
    for start in range(0, len(rows), batch_size):
        chunk = rows[start : start + batch_size]
        sku_values = {
            _as_text(row.get("sku")) for row in chunk if _as_text(row.get("sku"))
        }
        ean_values = {
            _as_text(row.get("ean")) for row in chunk if _as_text(row.get("ean"))
        }
        reference_values = {
            _as_text(row.get("product_reference"))
            for row in chunk
            if _as_text(row.get("product_reference"))
        }

        conditions = []
        if sku_values:
            conditions.append(Product.sku.in_(sku_values))
        if ean_values:
            conditions.append(Product.ean.in_(ean_values))
        if reference_values:
            conditions.append(Product.product_reference.in_(reference_values))

        existing_by_sku: dict[str, Product] = {}
        existing_by_ean: dict[str, Product] = {}
        existing_by_reference: dict[str, Product] = {}

        if conditions:
            existing_result = await db.execute(select(Product).where(or_(*conditions)))
            existing_products = existing_result.scalars().all()
            for product in existing_products:
                if product.sku:
                    existing_by_sku[product.sku] = product
                if product.ean:
                    existing_by_ean[product.ean] = product
                if product.product_reference:
                    existing_by_reference[product.product_reference] = product

        for row in chunk:
            sku = _as_text(row.get("sku"))
            ean = _as_text(row.get("ean"))
            product_reference = _as_text(row.get("product_reference"))

            matched_product: Product | None = None
            for candidate in (
                existing_by_sku.get(sku) if sku else None,
                existing_by_ean.get(ean) if ean else None,
                (
                    existing_by_reference.get(product_reference)
                    if product_reference
                    else None
                ),
            ):
                if candidate is not None:
                    matched_product = candidate
                    break

            if matched_product is None:
                matched_product = Product(**row)
                db.add(matched_product)
            else:
                for column in XLSX_COLUMN_MAP.values():
                    setattr(matched_product, column, row.get(column))

            await db.flush()

            if matched_product.sku:
                existing_by_sku[matched_product.sku] = matched_product
            if matched_product.ean:
                existing_by_ean[matched_product.ean] = matched_product
            if matched_product.product_reference:
                existing_by_reference[matched_product.product_reference] = (
                    matched_product
                )

        upserted_rows += len(chunk)
        if progress_callback is not None:
            await progress_callback(upserted_rows)

    await db.commit()
    return upserted_rows


async def _run_product_import_task(
    *,
    task_id: str,
    file_name: str,
    raw: bytes,
) -> None:
    async with SessionLocal() as session:
        task = await session.get(ProductImportTask, task_id)
        if task is None:
            return

        task.status = "running"
        task.started_at = datetime.utcnow()
        task.error_message = None
        await session.commit()

        try:
            parsed_rows = _read_xlsx_rows(raw)
            rows, skipped_rows = _deduplicate_rows(parsed_rows)
            task.total_rows = len(parsed_rows)
            task.skipped_rows = skipped_rows
            task.file_name = file_name
            await session.commit()

            async def update_progress(processed_rows: int) -> None:
                task.processed_rows = processed_rows
                task.upserted_rows = processed_rows
                await session.commit()
                await sleep(0)

            if rows:
                upserted_rows = await _upsert_products_in_batches(
                    session,
                    rows,
                    progress_callback=update_progress,
                )
            else:
                upserted_rows = 0

            task.status = "completed"
            task.processed_rows = len(parsed_rows)
            task.upserted_rows = upserted_rows
            task.finished_at = datetime.utcnow()
            task.error_message = None
            await session.commit()
        except Exception as exc:
            await session.rollback()
            task = await session.get(ProductImportTask, task_id)
            if task is None:
                return
            task.status = "failed"
            task.error_message = _summarize_task_error(exc)
            task.finished_at = datetime.utcnow()
            await session.commit()


def _is_all_categories_value(value: str | None) -> bool:
    if value is None:
        return True
    normalized = value.strip().lower()
    return normalized in {"", "all", "all categories", "all category", "allcategories"}


def _normalized_product_category_expression():
    return func.lower(func.trim(Product.product_category))


@router.get("/db")
async def get_db_products(
    db: AsyncSession = Depends(get_db),
    product_reference: Optional[str] = Query(None, alias="productReference"),
    page: int = Query(0, ge=0),
    sku: Optional[str] = Query(None),
    limit: int = Query(30, ge=1, le=1000),
    category: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    sort_by: str = Query("id", alias="sortBy"),
    sort_order: SortOrderEnum = Query(default=SortOrderEnum.DESC, alias="sortOrder"),
):
    """Return paginated spreadsheet-imported rows from the local DB."""
    sort_columns = {
        "id": Product.id,
        "sku": Product.sku,
        "productReference": Product.product_reference,
        "category": Product.product_category,
        "ean": Product.ean,
        "moin": Product.moin,
        "price": Product.price,
        "marketplaceStatus": Product.marketplace_status,
        "lastChangedAt": Product.last_changed_at,
    }
    sort_column = sort_columns.get(sort_by, Product.id)
    sorter = asc if sort_order == SortOrderEnum.ASC else desc

    filters = []
    if sku:
        filters.append(Product.sku == sku)
    if product_reference:
        filters.append(Product.product_reference == product_reference)
    if category and not _is_all_categories_value(category):
        normalized_category = category.strip().casefold()
        if normalized_category:
            filters.append(
                _normalized_product_category_expression() == normalized_category
            )
    if search:
        if term := search.strip():
            pattern = f"%{term}%"
            filters.append(
                or_(
                    Product.sku.ilike(pattern),
                    Product.product_reference.ilike(pattern),
                    Product.ean.ilike(pattern),
                    Product.moin.ilike(pattern),
                    Product.product_category.ilike(pattern),
                    Product.marketplace_status.ilike(pattern),
                    Product.error_message.ilike(pattern),
                    Product.active_status.ilike(pattern),
                )
            )

    count_stmt = select(func.count()).select_from(Product)
    if filters:
        count_stmt = count_stmt.where(*filters)
    total = await db.scalar(count_stmt)

    stmt = (
        select(Product)
        .order_by(sorter(sort_column), sorter(Product.id))
        .offset(page * limit)
        .limit(limit)
    )
    if filters:
        stmt = stmt.where(*filters)

    result = await db.execute(stmt)
    items = result.scalars().all()

    return {
        "items": [_product_to_dict(item) for item in items],
        "page": page,
        "limit": limit,
        "total": total or 0,
    }


@router.get("/db/categories")
async def get_db_product_categories(
    db: AsyncSession = Depends(get_db),
):
    """Return OTTO subcategories from the local category cache."""
    stmt = (
        select(func.trim(Category.name))
        .where(Category.name.is_not(None))
        .distinct()
        .order_by(func.trim(Category.name).asc())
    )
    result = await db.execute(stmt)
    raw_items = result.scalars().all()

    unique_items: list[str] = []
    seen: set[str] = set()
    for item in raw_items:
        if item is None:
            continue
        normalized = item.strip()
        if not normalized:
            continue
        key = normalized.casefold()
        if key in seen:
            continue
        seen.add(key)
        unique_items.append(normalized)

    unique_items.sort(key=str.casefold)
    return {
        "items": unique_items,
        "total": len(unique_items),
    }


@router.get("/db/category-groups/categories")
async def get_db_category_group_categories(
    category_group: list[str] = Query(default=[]),
    db: AsyncSession = Depends(get_db),
):
    """Return local OTTO subcategories only for the requested CategoryGroups."""
    requested_groups: list[str] = []
    seen_groups: set[str] = set()
    for item in category_group:
        normalized = str(item or "").strip()
        if not normalized:
            continue
        key = normalized.casefold()
        if key in seen_groups:
            continue
        seen_groups.add(key)
        requested_groups.append(normalized)

    if not requested_groups:
        return {"items": [], "total": 0}

    stmt = (
        select(CategoryGroup)
        .options(selectinload(CategoryGroup.categories))
        .where(func.lower(CategoryGroup.name).in_([item.casefold() for item in requested_groups]))
        .order_by(CategoryGroup.name.asc())
    )
    result = await db.execute(stmt)
    groups = result.scalars().unique().all()

    items: list[dict[str, Any]] = []
    for group in groups:
        category_items = sorted(
            [
                {
                    "name": str(category.name).strip(),
                    "nameRu": str(category.name_ru).strip() if category.name_ru else None,
                    "displayName": str(category.name).strip(),
                }
                for category in group.categories
                if category.name and str(category.name).strip()
            ],
            key=lambda item: item["name"].casefold(),
        )
        categories = [item["name"] for item in category_items]
        if not categories:
            categories = [str(group.name).strip()]
            category_items = [
                {
                    "name": str(group.name).strip(),
                    "nameRu": str(group.name_ru).strip() if group.name_ru else None,
                    "displayName": str(group.name).strip(),
                }
            ]
        items.append(
            {
                "categoryGroup": group.name,
                "categoryGroupRu": group.name_ru,
                "displayCategoryGroup": group.name,
                "categories": categories,
                "categoriesDisplay": category_items,
            }
        )

    return {"items": items, "total": len(items)}


@router.get("/db/category-attributes")
async def get_db_category_attributes(
    category: str | None = Query(default=None),
    category_group: str | None = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    """Return local OTTO attributes for a product category or CategoryGroup."""
    normalized_category = str(category or "").strip()
    normalized_group = str(category_group or "").strip()
    if not normalized_category and not normalized_group:
        return {"items": [], "total": 0, "categoryGroup": None}

    stmt = (
        select(CategoryGroup)
        .options(selectinload(CategoryGroup.attributes).selectinload(Attribute.allowed_values))
        .order_by(CategoryGroup.name.asc())
    )
    if normalized_category:
        stmt = stmt.join(Category).where(func.lower(Category.name) == normalized_category.casefold())
    else:
        stmt = stmt.where(func.lower(CategoryGroup.name) == normalized_group.casefold())

    result = await db.execute(stmt)
    group = result.scalars().unique().first()
    if group is None:
        return {"items": [], "total": 0, "categoryGroup": None}

    variation_attribute_ids = set(
        (
            await db.scalars(
                select(VariationTheme.attribute_id).where(
                    VariationTheme.group_id == group.id
                )
            )
        ).all()
    )
    attribute_fallbacks, value_fallbacks = await _load_attribute_translation_fallbacks_for_names(
        db,
        [attr.name for attr in group.attributes if attr.name],
    )

    items = [
        _attribute_payload(
            attr,
            attribute_fallbacks,
            value_fallbacks,
            is_variation_theme=attr.id in variation_attribute_ids,
        )
        for attr in sorted(group.attributes, key=lambda item: item.name.casefold())
        if attr.name
    ]
    return {
        "items": items,
        "total": len(items),
        "categoryGroup": group.name,
        "categoryGroupRu": group.name_ru,
        "displayCategoryGroup": group.name,
    }


@router.post("/{product_id}/variants/preview")
async def preview_product_variants(
    product_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Preview variation combinations without creating missing variants."""
    try:
        return await ProductVariantService(db).preview(product_id)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))


@router.post("/{product_id}/variants/generate")
async def generate_product_variants(
    product_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Create missing product variants for the current product attributes."""
    try:
        result = await ProductVariantService(db).generate(product_id)
        queued = 0
        queue_errors: list[str] = []
        for item in result.get("items", []):
            if not isinstance(item, dict):
                continue
            if item.get("source") == "source" or item.get("status") != "pending_generation":
                continue
            variant_id = item.get("id")
            if not isinstance(variant_id, int):
                continue
            try:
                await enqueue_job(
                    "regenerate_product_variant_image_task",
                    variant_id=variant_id,
                )
                queued += 1
            except Exception as exc:
                queue_errors.append(f"variant {variant_id}: {exc}")
        result["imageGenerationQueued"] = queued
        result["imageGenerationQueueErrors"] = queue_errors
        return result
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))


@router.get("/{product_id}/variants")
async def list_product_variants(
    product_id: int,
    db: AsyncSession = Depends(get_db),
):
    try:
        return await ProductVariantService(db).list_variants(product_id)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))


@router.patch("/{product_id}/variants/{variant_id}")
async def update_product_variant(
    product_id: int,
    variant_id: int,
    payload: dict[str, Any] | None = Body(default=None),
    db: AsyncSession = Depends(get_db),
):
    try:
        return await ProductVariantService(db).update_variant(
            product_id,
            variant_id,
            payload or {},
        )
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))


@router.delete("/{product_id}/variants/{variant_id}")
async def delete_product_variant(
    product_id: int,
    variant_id: int,
    db: AsyncSession = Depends(get_db),
):
    try:
        return await ProductVariantService(db).delete_variant(product_id, variant_id)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))


@router.post("/{product_id}/variants/{variant_id}/regenerate-image")
async def regenerate_product_variant_image(
    product_id: int,
    variant_id: int,
    db: AsyncSession = Depends(get_db),
):
    try:
        variant = await ProductVariantService(db).mark_image_regeneration_queued(
            product_id,
            variant_id,
        )
        await enqueue_job(
            "regenerate_product_variant_image_task",
            variant_id=variant_id,
        )
        return {"success": True, "queued": True, "variant": variant}
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(exc))
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Could not queue image regeneration: {exc}",
        )


@router.post("/variant-image/generate")
async def generate_product_variant_image(
    payload: dict[str, Any] = Body(default_factory=dict),
):
    combination_payload = payload.get("combination")
    if not isinstance(combination_payload, list):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="combination is required")

    combination: list[dict[str, str]] = []
    for item in combination_payload:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name")).strip()
        value = str(item.get("value")).strip()
        attribute_id = str(item.get("attributeId")).strip()
        if name and value:
            combination.append({"attributeId": attribute_id, "name": name, "value": value})
    if not combination:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="combination has no usable values")

    try:
        result = await generate_variant_image_from_snapshot(
            combination=combination,
            source_image_url=str(payload.get("sourceImageUrl") or "").strip() or None,
            request_id=str(payload.get("requestId") or "").strip() or None,
        )
        return {"success": True, **result}
    except httpx.HTTPStatusError as exc:
        detail = exc.response.text[:500] if exc.response is not None else str(exc)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail=f"Image provider error: {detail}",
        )
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=f"Could not generate image: {exc}",
        )


async def _get_otto_product_categories(
    *,
    page: int,
    limit: int,
    category: str | None,
    product_service: ProductService,
):
    payload = CategoryQuery(
        page=page,
        limit=limit,
        category=category,
    ).to_payload()
    return await product_service.get_categories(payload)


@router.get("/categories")
async def get_product_categories(
    page: int = Query(default=0, ge=0),
    limit: int = Query(default=1, ge=0, le=2000),
    category: str | None = Query(default=None),
    product_service: ProductService = Depends(get_product_service),
):
    """Proxy OTTO `GET /v5/products/categories` through the backend API."""
    return await _get_otto_product_categories(
        page=page,
        limit=limit,
        category=category,
        product_service=product_service,
    )


@otto_v5_router.get("/categories")
async def get_otto_v5_product_categories(
    page: int = Query(default=0, ge=0),
    limit: int = Query(default=10, ge=10, le=2000),
    category: str | None = Query(default=None),
    controller: Controller = Query(default=Controller.JV),
    product_service: ProductService = Depends(get_product_service),
):
    """Proxy OTTO `GET /v5/products/categories` and expose the OTTO path in docs."""
    return await _get_otto_product_categories(
        page=page,
        limit=limit,
        category=category,
        controller=controller,
        product_service=product_service,
    )


@router.get(
    "/otto",
    response_model=ProductResponse,
    summary="Fetch products directly from OTTO",
)
@otto_v5_router.get(
    "",
    response_model=ProductResponse,
    summary="Fetch products directly from OTTO",
)
async def get_otto_products_direct(
    controller: Controller = Query(default=Controller.JV),
    page: int = Query(default=0, ge=0),
    limit: int = Query(default=30, ge=1, le=1000),
    sku: str | None = Query(default=None),
    product_reference: str | None = Query(default=None, alias="productReference"),
    ean: str | None = Query(default=None),
    moin: str | None = Query(default=None),
    product_service: ProductService = Depends(get_product_service),
):
    """Proxy OTTO `GET /v5/products` directly through the backend API."""
    payload = {
        "page": page,
        "limit": limit,
        **({"sku": sku} if sku else {}),
        **({"productReference": product_reference} if product_reference else {}),
        **({"ean": ean} if ean else {}),
        **({"moin": moin} if moin else {}),
    }
    return await product_service.get_products(payload, controller=controller)


@router.get("/otto/shipping-profiles")
async def get_shipping_profiles(
    controller: Controller = Query(default=Controller.JV),
    product_service: ProductService = Depends(get_product_service),
):
    return await product_service.get_shipping_profiles(controller=controller)


# <======= POST METHOD =======>


@router.post("/create", response_model=ProductCreateResponse)
async def create_or_update_products(
    payload: CreateProductRequest,
    product_service: ProductService = Depends(get_product_service),
):
    """Create or update products in OTTO from already validated request payloads."""
    return await product_service.create_or_update_products(payload)


@router.post("/create-availability", response_model=AvailabilityResponse)
async def create_availability(
    payload: Availability,
    product_service: ProductService = Depends(get_product_service),
):
    MAPPER_LOGGER.info(
        "step=create_availability_endpoint_start sku=%s quantity=%s shipping_profile_id=%s controller=%s",
        payload.sku,
        payload.quantity,
        payload.shippingProfileID,
        payload.controller.value
        if isinstance(payload.controller, Controller)
        else payload.controller,
    )
    return await product_service.create_availability(payload)


@router.post("/deactivate-by-ean")
async def deactivate_products_by_ean(
    payload: dict[str, Any],
    db: AsyncSession = Depends(get_db),
    _current_user=Depends(require_role([RoleEnum.SEO])),
    product_service: ProductService = Depends(get_product_service),
):
    raw_eans = payload.get("eans")
    if not isinstance(raw_eans, list) or not raw_eans:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "eans must be a non-empty array"},
        )

    controller_value = str(payload.get("controller") or "jv").lower()
    try:
        controller = Controller(controller_value)
    except ValueError:
        return JSONResponse(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            content={
                "success": False,
                "message": f"Invalid controller: {controller_value}",
            },
        )

    eans = _normalize_ean_lines(raw_eans)
    if not eans:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "No valid EAN values provided"},
        )

    stmt = select(Product).where(Product.ean.in_(eans))
    existing_rows = (await db.execute(stmt)).scalars().all()
    products_by_ean = {str(item.ean): item for item in existing_rows if item.ean}

    semaphore = asyncio.Semaphore(PRODUCT_DEACTIVATE_CONCURRENCY)
    results: list[dict[str, Any] | None] = [None] * len(eans)

    async def _deactivate_one(index: int, ean: str) -> None:
        product_row = products_by_ean.get(ean)
        sku = str((product_row.sku if product_row and product_row.sku else ean)).strip()
        item_result = {
            "ean": ean,
            "sku": sku,
            "quantity_success": False,
            "status_success": False,
            "success": False,
            "message": "",
        }
        try:
            async with semaphore:
                quantity_response = await product_service.update_quantity(
                    {"sku": sku, "quantity": "0"},
                    controller=controller,
                )
                await product_service.update_status(
                    {"status": [{"sku": sku, "active": False}]},
                    controller=controller,
                )
            item_result["quantity_success"] = True
            item_result["status_success"] = True
            item_result["success"] = True
            item_result["message"] = "deactivated"
            if product_row is not None:
                product_row.active_status = "false"
                product_row.marketplace_status = "INACTIVE"
                product_row.error_message = None
            _ = quantity_response
        except Exception as exc:
            item_result["message"] = str(exc)
        results[index] = item_result

    await asyncio.gather(
        *[_deactivate_one(index, ean) for index, ean in enumerate(eans)]
    )
    await db.commit()

    normalized_results = [item for item in results if isinstance(item, dict)]
    failed = [item for item in normalized_results if not item.get("success")]
    return {
        "success": len(failed) == 0,
        "controller": controller.value,
        "total": len(eans),
        "failed": len(failed),
        "items": normalized_results,
    }


@router.post(
    "/upload-xlsx-task",
    response_model=ProductImportTaskDTO,
    responses={
        400: {"model": ProductCreationErrorResponse, "description": "Invalid request"},
        415: {
            "model": ProductCreationErrorResponse,
            "description": "Unsupported media type",
        },
    },
)
async def create_xlsx_import_task(
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
    file: UploadFile = File(..., description="XLSX file exported from OTTO market"),
):
    """Create a background XLSX import task and return its initial status."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        return JSONResponse(
            status_code=status.HTTP_415_UNSUPPORTED_MEDIA_TYPE,
            content=ProductCreationErrorResponse(
                message="Only .xlsx files are supported"
            ).model_dump(),
        )

    raw = await file.read()
    if not raw:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content=ProductCreationErrorResponse(
                message="Uploaded file is empty"
            ).model_dump(),
        )

    task = ProductImportTask(
        id=str(uuid4()),
        file_name=file.filename,
        status="queued",
        created_by_user_id=current_user.id,
        total_rows=None,
        processed_rows=0,
        upserted_rows=0,
        skipped_rows=0,
        error_message=None,
    )
    db.add(task)
    await db.commit()
    await db.refresh(task)

    background_tasks.add_task(
        _run_product_import_task,
        task_id=task.id,
        file_name=file.filename,
        raw=raw,
    )
    return _task_to_dto(task)


@router.get(
    "/import-tasks",
    response_model=ProductImportTaskListResponse,
)
async def list_product_import_tasks(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
    limit: int = Query(default=20, ge=1, le=100),
):
    stmt = (
        select(ProductImportTask)
        .where(ProductImportTask.created_by_user_id == current_user.id)
        .order_by(ProductImportTask.created_at.desc())
        .limit(limit)
    )
    result = await db.execute(stmt)
    tasks = result.scalars().all()
    return ProductImportTaskListResponse(items=[_task_to_dto(task) for task in tasks])


@router.get(
    "/import-tasks/{task_id}",
    response_model=ProductImportTaskDTO,
)
async def get_product_import_task(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    task = await db.get(ProductImportTask, task_id)
    if task is None or task.created_by_user_id != current_user.id:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content=ProductCreationErrorResponse(
                message=f"Import task '{task_id}' not found"
            ).model_dump(),
        )
    return _task_to_dto(task)


@router.post(
    "/ean-pool/import",
    response_model=EanPoolImportResponse,
    summary="Import EANs into the free pool",
)
async def import_ean_pool_items(
    payload: EanPoolImportRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    result = await EanPoolService(db).import_eans(
        payload.eans,
        source=payload.source,
        note=payload.note,
        created_by_user_id=current_user.id,
        metadata=payload.metadata,
    )
    return EanPoolImportResponse(**result)


@router.get(
    "/ean-pool/stats",
    response_model=EanPoolStatsResponse,
    summary="Get EAN pool counters",
)
async def get_ean_pool_stats(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    _ = current_user
    stats = await EanPoolService(db).stats()
    return EanPoolStatsResponse(**stats)


@router.get(
    "/ean-pool",
    response_model=EanPoolListResponse,
    summary="List EAN pool items",
)
async def list_ean_pool_items(
    status_filter: EanPoolStatus | None = Query(default=None, alias="status"),
    limit: int = Query(default=100, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    _ = current_user
    items = await EanPoolService(db).list_items(
        status=status_filter.value if status_filter else None,
        limit=limit,
    )
    return EanPoolListResponse(items=items)


@router.post(
    "/ean-pool/reserve",
    response_model=EanPoolItemResponse,
    summary="Reserve the next available EAN",
)
async def reserve_ean_pool_item(
    payload: EanPoolReserveRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    _ = current_user
    try:
        item = await EanPoolService(db).reserve_next(
            reserved_for=payload.reserved_for,
            metadata=payload.metadata,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=str(exc),
        ) from exc
    return EanPoolItemResponse(item=ean_pool_item_to_dict(item))


@router.post(
    "/ean-pool/{ean}/used",
    response_model=EanPoolItemResponse,
    summary="Mark an EAN as used",
)
async def mark_ean_pool_item_used(
    ean: str,
    payload: EanPoolMarkUsedRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    _ = current_user
    try:
        item = await EanPoolService(db).mark_used(ean, used_for=payload.used_for)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(exc),
        ) from exc
    return EanPoolItemResponse(item=ean_pool_item_to_dict(item))


@router.post(
    "/ean-pool/{ean}/release",
    response_model=EanPoolItemResponse,
    summary="Release a reserved EAN",
)
async def release_ean_pool_item(
    ean: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    _ = current_user
    try:
        item = await EanPoolService(db).release(ean)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        ) from exc
    return EanPoolItemResponse(item=ean_pool_item_to_dict(item))


@router.get("/db/{sku}")
async def get_db_product(
    sku: str,
    db: AsyncSession = Depends(get_db),
):
    """Fetch one product from the local DB by SKU (generic product lookup path)."""
    stmt = select(Product).where(Product.sku == sku)
    stmt = stmt.order_by(Product.id.desc())

    result = await db.execute(stmt)
    product = result.scalars().first()
    if not product:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={"message": f"Product with sku '{sku}' not found in DB"},
        )
    return _product_to_dict(product)


@router.post(
    "/tasks/create-from-factory", response_model=ProductFactoryCreateResponseDTO
)
async def create_product_task_from_factory(
    payload: ProductFactoryCreateRequestDTO,
    afterbuy: AfterbuyService = Depends(get_afterbuy_login),
    current_user=Depends(require_role([RoleEnum.SEO])),
    product_service: ProductService = Depends(get_product_service),
):
    run_id = payload.run_id or str(uuid4())
    now = datetime.now(UTC).isoformat()
    task = {
        "status": "IN_PROGRESS",
        "controller": payload.controller.value,
        "factory_id": payload.factory_id,
        "source_items": 0,
        "mapped_items": 0,
        "payload_items": 0,
        "issues": [],
        "products": [],
        "current_step": "prepare_queued",
        "current_step_started_at": now,
        "updated_at": now,
        "heartbeat_at": now,
        "progress_total": 0,
        "progress_completed": 0,
        "progress_percent": 0,
    }
    await _save_factory_task_state(
        run_id,
        task,
        created_by_user_id=current_user.id,
    )

    await enqueue_job(
        "prepare_factory_products_task",
        process_id=run_id,
        payload=payload.model_dump(mode="json"),
    )

    return ProductFactoryCreateResponseDTO(
        success=True,
        run_id=run_id,
        controller=payload.controller,
        factory_id=payload.factory_id,
        source_items=0,
        mapped_items=0,
        payload_items=0,
        issues=[],
        process_id=run_id,
        process_state="IN_PROGRESS",
    )


@router.get("/tasks/create-from-factory/latest")
async def get_latest_factory_prepare_task(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    stmt = (
        select(FactoryTaskState)
        .where(FactoryTaskState.created_by_user_id == current_user.id)
        .order_by(FactoryTaskState.updated_at.desc())
        .limit(20)
    )
    records = list((await db.execute(stmt)).scalars().all())
    if not records:
        legacy_stmt = (
            select(FactoryTaskState)
            .where(FactoryTaskState.created_by_user_id.is_(None))
            .order_by(FactoryTaskState.updated_at.desc())
            .limit(20)
        )
        records = list((await db.execute(legacy_stmt)).scalars().all())

    for record in records:
        process_id = record.process_id
        task = await _get_owned_factory_task_state(process_id, current_user.id)
        if task is None:
            continue
        (
            task,
            heartbeat_lag_sec,
            step_elapsed_sec,
            is_stuck,
        ) = await _mark_factory_task_stale_if_needed(process_id, task)
        if _is_empty_stale_factory_failure(task):
            continue
        current_step = str(task.get("current_step") or "")
        task_has_products = _has_factory_task_products(task)
        if task.get("status") == "DONE" and current_step in FACTORY_FINAL_STEPS:
            continue
        if (
            task.get("status") == "FAILED"
            and current_step in FACTORY_FINAL_STEPS
            and not task_has_products
        ):
            continue
        return {
            "success": True,
            "process_id": process_id,
            "process_state": task.get("status"),
            "heartbeat_lag_sec": heartbeat_lag_sec,
            "step_elapsed_sec": step_elapsed_sec,
            "stuck": is_stuck or bool(task.get("stuck")),
            "stuck_message": task.get("stuck_message"),
            **task,
        }

    return JSONResponse(
        status_code=status.HTTP_404_NOT_FOUND,
        content={"success": False, "message": "No saved creation draft"},
    )


@router.delete("/tasks/create-from-factory")
async def delete_factory_prepare_tasks(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    stmt = select(FactoryTaskState.process_id).where(
        or_(
            FactoryTaskState.created_by_user_id == current_user.id,
            FactoryTaskState.created_by_user_id.is_(None),
        )
    )
    process_ids = [str(item) for item in (await db.execute(stmt)).scalars().all()]

    for process_id in process_ids:
        FACTORY_PREPARE_TASKS.pop(process_id, None)
        await FACTORY_TASK_STATE_SERVICE.delete_task(process_id)

    return {
        "success": True,
        "deleted": len(process_ids),
    }


@router.get("/tasks/create-from-factory/{process_id}")
async def get_factory_prepare_task(
    process_id: str,
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    task = await _get_owned_factory_task_state(process_id, current_user.id)
    if task is None:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "success": False,
                "message": "Task not found",
                "process_id": process_id,
            },
        )
    (
        task,
        heartbeat_lag_sec,
        step_elapsed_sec,
        is_stuck,
    ) = await _mark_factory_task_stale_if_needed(process_id, task)

    return {
        "success": True,
        "process_id": process_id,
        "process_state": task.get("status"),
        "heartbeat_lag_sec": heartbeat_lag_sec,
        "step_elapsed_sec": step_elapsed_sec,
        "stuck": is_stuck or bool(task.get("stuck")),
        "stuck_message": task.get("stuck_message"),
        **task,
    }


@router.delete("/tasks/create-from-factory/{process_id}")
async def delete_factory_prepare_task(
    process_id: str,
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    task = await _get_owned_factory_task_state(process_id, current_user.id)
    if task is None:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "success": False,
                "message": "Task not found",
                "process_id": process_id,
            },
        )
    FACTORY_PREPARE_TASKS.pop(process_id, None)
    await FACTORY_TASK_STATE_SERVICE.delete_task(process_id)
    return {
        "success": True,
        "process_id": process_id,
    }


@router.patch("/tasks/create-from-factory/{process_id}/draft")
async def save_factory_prepare_task_draft(
    process_id: str,
    payload: dict[str, Any],
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    task = await _get_owned_factory_task_state(process_id, current_user.id)
    if task is None:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "success": False,
                "message": "Task not found",
                "process_id": process_id,
            },
        )

    task_step = str(task.get("current_step") or "")
    task_status = str(task.get("status") or "")
    is_enrichment_running = task_status == "IN_PROGRESS" and task_step.startswith(
        "ai_enrichment"
    )

    products = payload.get("products")
    if isinstance(products, list) and not is_enrichment_running:
        task["products"] = products

    current_step = payload.get("current_step")
    if (
        isinstance(current_step, str)
        and current_step.strip()
        and not is_enrichment_running
    ):
        task["current_step"] = current_step.strip()

    frontend_draft_payload = payload.get("frontend_draft")
    frontend_draft = (
        dict(task.get("frontend_draft"))
        if isinstance(task.get("frontend_draft"), dict)
        else {}
    )
    if isinstance(frontend_draft_payload, dict):
        if is_enrichment_running:
            frontend_draft_payload = {
                key: value
                for key, value in frontend_draft_payload.items()
                if key != "workflowStep"
            }
        frontend_draft.update(frontend_draft_payload)
    task["frontend_draft"] = frontend_draft
    task["updated_at"] = datetime.now(UTC).isoformat()

    await _save_factory_task_state(
        process_id,
        task,
        created_by_user_id=current_user.id,
    )
    return {
        "success": True,
        "process_id": process_id,
        "updated_at": task.get("updated_at"),
    }


@router.websocket("/tasks/create-from-factory/{process_id}/ws")
async def factory_prepare_task_ws(websocket: WebSocket, process_id: str):
    await websocket.accept()
    current_user = await _get_websocket_user(websocket)
    if current_user is None:
        await websocket.send_json(
            {
                "success": False,
                "process_id": process_id,
                "message": "Unauthorized",
            }
        )
        await websocket.close(code=4401)
        return
    try:
        while True:
            task = await _get_owned_factory_task_state(process_id, current_user.id)
            if task is None:
                await websocket.send_json(
                    {
                        "success": False,
                        "process_id": process_id,
                        "message": "Task not found",
                    }
                )
                await websocket.close(code=4404)
                return
            (
                task,
                heartbeat_lag_sec,
                step_elapsed_sec,
                _is_stuck,
            ) = await _mark_factory_task_stale_if_needed(process_id, task)

            await websocket.send_json(
                {
                    "success": True,
                    "process_id": process_id,
                    "process_state": task.get("status"),
                    "heartbeat_lag_sec": heartbeat_lag_sec,
                    "step_elapsed_sec": step_elapsed_sec,
                    **task,
                }
            )

            if task.get("status") in {"DONE", "FAILED"}:
                await websocket.close(code=1000)
                return

            await asyncio.sleep(1)
    except WebSocketDisconnect:
        return


def _compact_source_specifics(source_item: dict[str, Any] | None) -> dict[str, str]:
    if not isinstance(source_item, dict):
        return {}
    xml_data = source_item.get("CustomItemSpecifics")
    if not isinstance(xml_data, str) or not xml_data.strip():
        return {}
    try:
        root = ET.fromstring(xml_data)
    except ET.ParseError:
        return {}

    result: dict[str, str] = {}
    for item in root.findall("NameValueList"):
        name = item.findtext("Name")
        if not name:
            continue
        values = [
            value.text.strip()
            for value in item.findall("Value")
            if value.text and value.text.strip()
        ]
        if values:
            result[name.strip()] = ", ".join(values)
    return result


def _normalized_identity_value(value: Any) -> str:
    return str(value or "").strip().casefold()


def _source_identity_values(source_item: dict[str, Any] | None) -> set[str]:
    if not isinstance(source_item, dict):
        return set()

    specifics = _compact_source_specifics(source_item)
    raw_values = [
        source_item.get("EAN"),
        source_item.get("ean"),
        source_item.get("SKU"),
        source_item.get("sku"),
        source_item.get("productReference"),
        source_item.get("product_reference"),
        specifics.get("EAN"),
        specifics.get("ean"),
    ]
    return {
        normalized
        for normalized in (_normalized_identity_value(value) for value in raw_values)
        if normalized
    }


def _product_identity_values(product: dict[str, Any] | None) -> set[str]:
    if not isinstance(product, dict):
        return set()
    raw_values = [
        product.get("ean"),
        product.get("EAN"),
        product.get("sku"),
        product.get("SKU"),
        product.get("productReference"),
        product.get("product_reference"),
    ]
    return {
        normalized
        for normalized in (_normalized_identity_value(value) for value in raw_values)
        if normalized
    }


def _build_source_item_lookup(
    source_items: list[Any],
) -> dict[str, dict[str, Any]]:
    lookup: dict[str, dict[str, Any]] = {}
    for source_item in source_items:
        if not isinstance(source_item, dict):
            continue
        for identity in _source_identity_values(source_item):
            lookup.setdefault(identity, source_item)
    return lookup


def _find_source_item_for_product(
    *,
    product: dict[str, Any],
    index: int,
    source_items: list[Any],
    source_items_by_identity: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    for identity in _product_identity_values(product):
        source_item = source_items_by_identity.get(identity)
        if source_item is not None:
            return source_item

    if index < len(source_items) and isinstance(source_items[index], dict):
        fallback = source_items[index]
        product_identities = _product_identity_values(product)
        fallback_identities = _source_identity_values(fallback)
        if not product_identities or product_identities.intersection(fallback_identities):
            return fallback

        MAPPER_LOGGER.warning(
            "step=source_item_index_mismatch index=%s product_ids=%s source_ids=%s",
            index,
            sorted(product_identities),
            sorted(fallback_identities),
        )

    return None


def _build_aftercool_comparison(
    *,
    source_item: dict[str, Any] | None,
    generated_product: dict[str, Any],
) -> dict[str, Any]:
    source = source_item if isinstance(source_item, dict) else {}
    specifics = _compact_source_specifics(source)
    description = (
        generated_product.get("productDescription")
        if isinstance(generated_product.get("productDescription"), dict)
        else {}
    )
    generated_attrs = description.get("attributes")
    aftercool_attrs = [
        {"name": key, "values": [value]}
        for key, value in sorted(specifics.items(), key=lambda item: item[0].casefold())
        if key and value
    ]

    return {
        "approved": False,
        "aftercool": {
            "title": source.get("Artikelbeschreibung"),
            "description": (
                source.get("StammartikelBeschreibungDetailsHtml")
                or source.get("TranslatedDescription")
                or source.get("Description")
                or source.get("Beschreibung")
            ),
            "attributes": aftercool_attrs,
            "price": source.get("Startpreis"),
            "ean": source.get("EAN") or source.get("ean") or specifics.get("EAN"),
        },
        "generated": {
            "title": description.get("productLine"),
            "description": description.get("description"),
            "bulletPoints": description.get("bulletPoints") or [],
            "attributes": generated_attrs if isinstance(generated_attrs, list) else [],
            "category": description.get("category"),
            "ean": generated_product.get("ean"),
        },
    }


async def _run_factory_enrichment_task(
    *,
    process_id: str,
    payload: dict[str, Any],
    product_service: ProductService,
) -> None:
    from app.mapper.product_mapper import ProductMapper

    task = await _get_factory_task_state(process_id)
    products = payload.get("products")
    if task is None or not isinstance(products, list) or not products:
        MAPPER_LOGGER.error(
            "Ошибка начала генерации: process_id=%s has_task=%s has_products=%s",
            process_id,
            task is not None,
            isinstance(products, list) and bool(products),
        )
        return

    source_items_raw = task.get("source_items_raw")
    source_items = source_items_raw if isinstance(source_items_raw, list) else []
    source_items_by_identity = _build_source_item_lookup(source_items)
    controller_value = str(
        task.get("controller") or payload.get("controller") or "jv"
    ).lower()
    controller = Controller(controller_value)
    category_group_contexts = await _load_category_group_contexts()
    ai_mapper = ProductMapper(
        products=[],
        controller=controller.value,
        otto_client=product_service.client,
        category_group_contexts=category_group_contexts,
    )

    now = datetime.now(UTC).isoformat()
    task["status"] = "IN_PROGRESS"
    task["current_step"] = "ai_enrichment_in_progress"
    task["current_step_started_at"] = now
    task["updated_at"] = now
    task["heartbeat_at"] = now
    task["stuck"] = False
    task["stuck_message"] = None
    task["progress_total"] = len(products)
    task["progress_completed"] = 0
    task["progress_percent"] = 0
    task["products"] = products
    task["partial_products_by_index"] = {}
    await _save_factory_task_state(process_id, task)

    MAPPER_LOGGER.info(
        "MAPPER начался успешно: process_id=%s products=%s",
        process_id,
        len(products),
    )

    heartbeat_stop = asyncio.Event()
    progress_lock = asyncio.Lock()

    async def _heartbeat_loop() -> None:
        while not heartbeat_stop.is_set():
            now_heartbeat = datetime.now(UTC).isoformat()
            async with progress_lock:
                if task.get("status") != "IN_PROGRESS":
                    return
                task["heartbeat_at"] = now_heartbeat
                task["updated_at"] = now_heartbeat
                task["heartbeat_count"] = int(task.get("heartbeat_count", 0)) + 1
                await _save_factory_task_state(process_id, task)
            await asyncio.sleep(FACTORY_HEARTBEAT_INTERVAL_SEC)

    heartbeat_task = asyncio.create_task(_heartbeat_loop())

    enriched_products: list[dict[str, Any] | None] = [None] * len(products)
    item_failures: list[str] = []
    work_queue: asyncio.Queue[tuple[int, dict[str, Any]] | None] = asyncio.Queue()

    try:

        async def _enrich_one(index: int, item: dict[str, Any]) -> None:
            model = ProductPayload.model_validate(item)
            model_dump = model.model_dump(mode="json", exclude_none=True)
            source_item = _find_source_item_for_product(
                product=model_dump,
                index=index,
                source_items=source_items,
                source_items_by_identity=source_items_by_identity,
            )
            MAPPER_LOGGER.info(
                "step=category_approval_enrichment_item_start process_id=%s index=%s sku=%s category=%s source_available=%s",
                process_id,
                index,
                model.sku,
                model.productDescription.category,
                source_item is not None,
            )
            enriched_payload = await asyncio.wait_for(
                ai_mapper.enrich_after_category_approval(
                    model_dump,
                    source_item=source_item,
                ),
                timeout=FACTORY_AI_ENRICH_ITEM_TIMEOUT_SEC,
            )
            enriched_model = ProductPayload.model_validate(enriched_payload)
            enriched_dump = enriched_model.model_dump(
                mode="json", exclude_none=True
            )
            for key, value in item.items():
                if key not in enriched_dump:
                    enriched_dump[key] = value
            enriched_dump["aftercoolComparison"] = _build_aftercool_comparison(
                source_item=source_item,
                generated_product=enriched_dump,
            )
            enriched_products[index] = enriched_dump
            async with progress_lock:
                by_index_raw = task.get("partial_products_by_index")
                by_index = by_index_raw if isinstance(by_index_raw, dict) else {}
                by_index[str(index)] = enriched_dump
                task["partial_products_by_index"] = by_index
                current_products = task.get("products")
                next_products = list(current_products) if isinstance(current_products, list) else list(products)
                if index < len(next_products):
                    next_products[index] = enriched_dump
                task["products"] = next_products
                task["progress_completed"] = int(task.get("progress_completed", 0)) + 1
                task["progress_percent"] = int(
                    round((task["progress_completed"] / max(1, len(products))) * 100)
                )
                progress_at = datetime.now(UTC).isoformat()
                task["updated_at"] = progress_at
                task["heartbeat_at"] = progress_at
                await _save_factory_task_state(process_id, task)
            MAPPER_LOGGER.info(
                "step=category_approval_enrichment_item_done process_id=%s index=%s sku=%s bullet_points=%s attributes=%s",
                process_id,
                index,
                enriched_model.sku,
                len(enriched_model.productDescription.bulletPoints),
                len(enriched_model.productDescription.attributes),
            )

        async def _worker(worker_id: int) -> None:
            while True:
                queued = await work_queue.get()
                try:
                    if queued is None:
                        return
                    index, item = queued
                    try:
                        await _enrich_one(index, item)
                    except Exception as exc:
                        fallback_model = ProductPayload.model_validate(item)
                        fallback_dump = fallback_model.model_dump(
                            mode="json",
                            exclude_none=True,
                        )
                        for key, value in item.items():
                            if key not in fallback_dump:
                                fallback_dump[key] = value
                        fallback_source_item = _find_source_item_for_product(
                            product=fallback_dump,
                            index=index,
                            source_items=source_items,
                            source_items_by_identity=source_items_by_identity,
                        )
                        fallback_dump["aftercoolComparison"] = _build_aftercool_comparison(
                            source_item=fallback_source_item,
                            generated_product=fallback_dump,
                        )
                        enriched_products[index] = fallback_dump
                        message = f"AI enrichment failed at index={index} sku={fallback_model.sku}: {exc}"
                        item_failures.append(message)
                        MAPPER_LOGGER.exception(
                            "step=category_approval_enrichment_item_failed process_id=%s worker=%s index=%s sku=%s error=%s",
                            process_id,
                            worker_id,
                            index,
                            fallback_model.sku,
                            exc,
                        )
                        async with progress_lock:
                            by_index_raw = task.get("partial_products_by_index")
                            by_index = by_index_raw if isinstance(by_index_raw, dict) else {}
                            by_index[str(index)] = fallback_dump
                            task["partial_products_by_index"] = by_index
                            current_products = task.get("products")
                            next_products = list(current_products) if isinstance(current_products, list) else list(products)
                            if index < len(next_products):
                                next_products[index] = fallback_dump
                            task["products"] = next_products
                            task["progress_completed"] = (
                                int(task.get("progress_completed", 0)) + 1
                            )
                            task["progress_percent"] = int(
                                round(
                                    (task["progress_completed"] / max(1, len(products)))
                                    * 100
                                )
                            )
                            progress_at = datetime.now(UTC).isoformat()
                            task["updated_at"] = progress_at
                            task["heartbeat_at"] = progress_at
                            await _save_factory_task_state(process_id, task)
                finally:
                    work_queue.task_done()

        for index, item in enumerate(products):
            await work_queue.put((index, item))

        worker_count = min(FACTORY_PRODUCT_CONCURRENCY, len(products))
        workers = [
            asyncio.create_task(_worker(worker_id)) for worker_id in range(worker_count)
        ]
        for _ in workers:
            await work_queue.put(None)
        await work_queue.join()
        await asyncio.gather(*workers)

        task["status"] = "DONE"
        task["products"] = [
            item for item in enriched_products if isinstance(item, dict)
        ]
        task["partial_products_by_index"] = {}
        task["current_step"] = "ai_enrichment_done"
        task["current_step_started_at"] = datetime.now(UTC).isoformat()
        task["updated_at"] = task["current_step_started_at"]
        task["heartbeat_at"] = task["current_step_started_at"]
        task["enriched_at"] = task["current_step_started_at"]
        task["finished_at"] = task["current_step_started_at"]
        task["progress_completed"] = len(products)
        task["progress_total"] = len(products)
        task["progress_percent"] = 100
        if item_failures:
            task["issues"] = [
                *item_failures[:100],
                *(
                    [f"{len(item_failures) - 100} more AI enrichment item failures"]
                    if len(item_failures) > 100
                    else []
                ),
            ]
        await _save_factory_task_state(process_id, task)
        MAPPER_LOGGER.info(
            "Маппер полностью закончен process_id=%s products=%s item_failures=%s",
            process_id,
            len(task["products"]),
            len(item_failures),
        )
    except Exception as exc:
        task["status"] = "FAILED"
        task["current_step"] = "ai_enrichment_failed"
        task["current_step_started_at"] = datetime.now(UTC).isoformat()
        task["updated_at"] = task["current_step_started_at"]
        task["heartbeat_at"] = task["current_step_started_at"]
        task["issues"] = [str(exc)]
        await _save_factory_task_state(process_id, task)
        MAPPER_LOGGER.exception(
            "Ошибка генерации process_id=%s error=%s",
            process_id,
            exc,
        )
    finally:
        heartbeat_stop.set()
        try:
            await heartbeat_task
        except Exception:
            pass


@router.post("/tasks/create-from-factory/{process_id}/enrich")
async def enrich_factory_prepared_products(
    process_id: str,
    payload: dict[str, Any],
    product_service: ProductService = Depends(get_product_service),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    task = await _get_owned_factory_task_state(process_id, current_user.id)
    MAPPER_LOGGER.info(
        "step=category_approval_enrichment_start process_id=%s has_task=%s",
        process_id,
        task is not None,
    )
    if task is None:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "success": False,
                "message": "Task not found",
                "process_id": process_id,
            },
        )
    products = payload.get("products")
    if not isinstance(products, list) or not products:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "products must be a non-empty array"},
        )

    controller_value = str(
        (task or {}).get("controller") or payload.get("controller") or "jv"
    ).lower()
    try:
        controller = Controller(controller_value)
    except ValueError:
        return JSONResponse(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            content={
                "success": False,
                "message": f"Invalid controller: {controller_value}",
                "process_id": process_id,
            },
        )

    if (
        task is not None
        and task.get("status") == "IN_PROGRESS"
        and task.get("current_step") == "ai_enrichment_in_progress"
    ):
        return {
            "success": True,
            "process_id": process_id,
            "process_state": "IN_PROGRESS",
            "queued": False,
        }

    if task is not None:
        now = datetime.now(UTC).isoformat()
        task["status"] = "IN_PROGRESS"
        task["current_step"] = "ai_enrichment_queued"
        task["current_step_started_at"] = now
        task["updated_at"] = now
        task["heartbeat_at"] = now
        task["stuck"] = False
        task["stuck_message"] = None
        task["progress_total"] = len(products)
        task["progress_completed"] = 0
        task["progress_percent"] = 0
        task["products"] = products
        task["partial_products_by_index"] = {}
        await _save_factory_task_state(
            process_id,
            task,
            created_by_user_id=current_user.id,
        )

    await enqueue_job(
        "enrich_factory_products_task",
        process_id=process_id,
        payload=payload,
    )

    return {
        "success": True,
        "process_id": process_id,
        "process_state": "IN_PROGRESS",
        "queued": True,
    }


def _contains_cyrillic(value: str) -> bool:
    return bool(re.search(r"[А-Яа-яЁё]", value or ""))


async def _load_attribute_allowed_value_lookup() -> dict[str, dict[str, str]]:
    async with SessionLocal() as session:
        result = await session.execute(
            select(Attribute)
            .options(selectinload(Attribute.allowed_values))
            .order_by(Attribute.name.asc())
        )
        lookup: dict[str, dict[str, str]] = {}
        for attr in result.scalars().unique().all():
            attr_key = normalize_translation_text(attr.name).casefold()
            if not attr_key:
                continue
            values: dict[str, str] = {}
            for item in attr.allowed_values:
                original = normalize_translation_text(item.value)
                if not original:
                    continue
                values[original.casefold()] = original
                if item.value_ru:
                    values[normalize_translation_text(item.value_ru).casefold()] = original
            if values:
                lookup[attr_key] = values
        return lookup


async def _translate_user_attribute_values_for_otto(
    products: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    allowed_value_lookup = await _load_attribute_allowed_value_lookup()
    async with SessionLocal() as session:
        translator = TranslationService(session)
        translated_products: list[dict[str, Any]] = []
        for product in products:
            next_product = dict(product)
            description = dict(next_product.get("productDescription") or {})
            attributes = description.get("attributes")
            if not isinstance(attributes, list):
                translated_products.append(next_product)
                continue

            translated_attributes: list[Any] = []
            for raw_attr in attributes:
                if not isinstance(raw_attr, dict):
                    translated_attributes.append(raw_attr)
                    continue
                attr = dict(raw_attr)
                attr_name = normalize_translation_text(str(attr.get("name") or ""))
                allowed_map = allowed_value_lookup.get(attr_name.casefold(), {})
                raw_values = attr.get("values")
                if not isinstance(raw_values, list):
                    translated_attributes.append(attr)
                    continue

                next_values: list[Any] = []
                for raw_value in raw_values:
                    if not isinstance(raw_value, str):
                        next_values.append(raw_value)
                        continue
                    normalized_value = normalize_translation_text(raw_value)
                    if not normalized_value:
                        next_values.append(raw_value)
                        continue

                    allowed_original = allowed_map.get(normalized_value.casefold())
                    if allowed_original:
                        next_values.append(allowed_original)
                        continue

                    if not _contains_cyrillic(normalized_value):
                        next_values.append(normalized_value)
                        continue

                    try:
                        next_values.append(
                            await translator.translate(
                                normalized_value,
                                source_lang="RU",
                                target_lang="DE",
                                context="user_attribute_input",
                            )
                        )
                    except Exception as exc:
                        MAPPER_LOGGER.warning(
                            "step=translate_user_attribute_failed attribute=%s value=%s error=%s",
                            attr_name,
                            normalized_value,
                            exc,
                        )
                        next_values.append(normalized_value)

                attr["values"] = next_values
                translated_attributes.append(attr)

            description["attributes"] = translated_attributes
            next_product["productDescription"] = description
            translated_products.append(next_product)
        return translated_products


OTTO_ERROR_TRANSLATION_FIELDS = {"title", "message", "description", "detail"}


def _extract_otto_error_issues(value: Any) -> list[str]:
    if not isinstance(value, dict):
        return []

    results = value.get("results")
    if not isinstance(results, list):
        return []

    issues: list[str] = []
    for entry in results:
        if not isinstance(entry, dict):
            continue
        variation = str(entry.get("variation") or "unknown").strip() or "unknown"
        errors = entry.get("errors")
        if not isinstance(errors, list):
            errors = [entry]
        for error in errors:
            if not isinstance(error, dict):
                continue
            code = str(error.get("code") or "error").strip() or "error"
            title = str(
                error.get("title")
                or error.get("message")
                or error.get("description")
                or error.get("detail")
                or ""
            ).strip()
            json_path = str(error.get("jsonPath") or error.get("path") or "").strip()
            parts = [variation, code]
            if title:
                parts.append(title)
            if json_path:
                parts.append(json_path)
            issues.append(": ".join(parts))
    return issues


async def _translate_otto_error_payload_for_ui(value: Any) -> Any:
    async with SessionLocal() as session:
        translator = TranslationService(session)

        async def translate_value(current: Any) -> Any:
            if isinstance(current, list):
                return [await translate_value(item) for item in current]
            if not isinstance(current, dict):
                return current

            translated: dict[str, Any] = {}
            for key, raw_value in current.items():
                if (
                    key in OTTO_ERROR_TRANSLATION_FIELDS
                    and isinstance(raw_value, str)
                    and normalize_translation_text(raw_value)
                ):
                    translated[f"{key}Original"] = raw_value
                    try:
                        translated[key] = await translator.translate(
                            raw_value,
                            source_lang="DE",
                            target_lang="RU",
                            context="otto_error",
                        )
                    except Exception as exc:
                        MAPPER_LOGGER.warning(
                            "step=translate_otto_error_failed field=%s error=%s",
                            key,
                            exc,
                        )
                        translated[key] = raw_value
                    continue

                translated[key] = await translate_value(raw_value)
            return translated

        return await translate_value(value)


def _variant_reserved_for(
    process_id: str,
    product: dict[str, Any],
    variant: dict[str, Any],
    index: int,
) -> str:
    product_reference = str(
        product.get("productReference") or product.get("sku") or index
    ).strip()
    combination_key = str(
        variant.get("combinationKey")
        or variant.get("combination_key")
        or variant.get("sku")
        or index
    ).strip()
    return f"factory-submit:{process_id}:{product_reference}:{combination_key}"


def _set_variant_ean(variant: dict[str, Any], ean: str) -> None:
    variant["ean"] = ean
    product_payload = variant.get("productPayload")
    if isinstance(product_payload, dict):
        product_payload["ean"] = ean


async def _fill_missing_variant_eans_from_pool(
    *,
    process_id: str,
    products: list[dict[str, Any]],
) -> list[dict[str, str]]:
    assignments: list[dict[str, str]] = []

    async with SessionLocal() as session:
        service = EanPoolService(session)
        for product_index, product in enumerate(products):
            variants = product.get("variants")
            if not isinstance(variants, list):
                continue
            for variant_index, variant in enumerate(variants):
                if not isinstance(variant, dict):
                    continue
                if variant.get("active") is False:
                    continue
                current_ean = str(variant.get("ean") or "").strip()
                if current_ean:
                    continue

                reserved_for = _variant_reserved_for(
                    process_id,
                    product,
                    variant,
                    variant_index,
                )
                item = await service.reserve_next(
                    reserved_for=reserved_for,
                    metadata={
                        "processId": process_id,
                        "productIndex": product_index,
                        "variantIndex": variant_index,
                    },
                )
                _set_variant_ean(variant, item.ean)
                assignments.append(
                    {
                        "ean": item.ean,
                        "reservedFor": reserved_for,
                        "productIndex": str(product_index),
                        "variantIndex": str(variant_index),
                    }
                )

    return assignments


async def _mark_ean_pool_assignments_used(assignments: list[dict[str, str]]) -> None:
    if not assignments:
        return
    async with SessionLocal() as session:
        service = EanPoolService(session)
        for assignment in assignments:
            ean = assignment.get("ean")
            used_for = assignment.get("reservedFor") or "factory-submit"
            if not ean:
                continue
            try:
                await service.mark_used(ean, used_for=used_for)
            except Exception as exc:
                MAPPER_LOGGER.warning(
                    "step=ean_pool_mark_used_failed ean=%s error=%s",
                    ean,
                    exc,
                )


async def _release_ean_pool_assignments(assignments: list[dict[str, str]]) -> None:
    if not assignments:
        return
    async with SessionLocal() as session:
        service = EanPoolService(session)
        for assignment in assignments:
            ean = assignment.get("ean")
            if not ean:
                continue
            try:
                await service.release(ean)
            except Exception as exc:
                MAPPER_LOGGER.warning(
                    "step=ean_pool_release_failed ean=%s error=%s",
                    ean,
                    exc,
                )


def _controller_compliance_payload(controller: Controller) -> dict[str, Any] | None:
    compliance = settings.compliance.get(controller)
    if compliance is None:
        return None
    return compliance.model_dump(mode="json", exclude_none=True)


def _apply_account_fields_to_product(
    product: dict[str, Any],
    *,
    controller: Controller,
    identity_value: str | None = None,
) -> dict[str, Any]:
    updated = dict(product)
    normalized_identity = str(identity_value or "").strip()
    if normalized_identity:
        updated["ean"] = normalized_identity
        updated["sku"] = normalized_identity
        updated["productReference"] = normalized_identity

    description = dict(updated.get("productDescription") or {})
    description["brandId"] = brand_id_for_controller(controller)
    updated["productDescription"] = description

    compliance = _controller_compliance_payload(controller)
    if compliance is not None:
        updated["compliance"] = compliance

    return updated


def _create_product_body(product: dict[str, Any]) -> dict[str, Any]:
    return {
        key: value
        for key, value in product.items()
        if key not in CREATE_PRODUCT_EXCLUDED_FIELDS
    }


def _availability_payload_item(
    *,
    controller: Controller,
    product: dict[str, Any],
) -> dict[str, str]:
    return {
        "controller": controller.value,
        "sku": str(product.get("sku") or "").strip(),
        "shippingProfileID": str(product.get("shippingProfileID") or "").strip(),
        "quantity": str(product.get("quantity") or "20").strip() or "20",
        "processingTime": str(product.get("processingTime") or "DEFAULT").strip()
        or "DEFAULT",
    }


def _clone_products_with_account_map(
    products: list[dict[str, Any]],
    ean_map: dict[str, str],
    *,
    controller: Controller,
) -> list[dict[str, Any]]:
    cloned: list[dict[str, Any]] = []
    for product in copy.deepcopy(products):
        current_ean = str(product.get("ean") or "").strip()
        mapped_ean = ean_map.get(current_ean)
        cloned.append(
            _apply_account_fields_to_product(
                product,
                controller=controller,
                identity_value=mapped_ean or current_ean,
            )
        )
    return cloned


async def _find_counterpart_factory_id(
    *,
    source_factory_id: str,
    target_controller: Controller,
) -> str | None:
    normalized_source_id = str(source_factory_id or "").strip()
    if not normalized_source_id:
        return None

    async with SessionLocal() as session:
        source_result = await session.execute(
            select(Factories).where(Factories.factory_id == normalized_source_id)
        )
        source_factory = source_result.scalar_one_or_none()
        if source_factory is None:
            MAPPER_LOGGER.warning(
                "step=xl_factory_lookup_missing_source factory_id=%s",
                normalized_source_id,
            )
            return None

        source_name = str(source_factory.name or "").strip()
        if not source_name:
            return None

        target_result = await session.execute(
            select(Factories)
            .where(
                func.lower(func.trim(Factories.name)) == source_name.lower(),
                func.upper(Factories.account) == target_controller.value.upper(),
            )
            .order_by(Factories.items_count.desc(), Factories.factory_id.asc())
        )
        target_factory = target_result.scalars().first()
        if target_factory is None:
            MAPPER_LOGGER.warning(
                "step=xl_factory_lookup_not_found source_factory_id=%s source_name=%s target_account=%s",
                normalized_source_id,
                source_name,
                target_controller.value,
            )
            return None

        return str(target_factory.factory_id)


async def _map_eans_from_counterpart_factory_by_index(
    *,
    afterbuy: AfterbuyService,
    source_factory_id: str,
    target_controller: Controller,
    source_items: list[Any],
    products: list[dict[str, Any]],
) -> tuple[dict[str, str], dict[str, Any]]:
    target_factory_id = await _find_counterpart_factory_id(
        source_factory_id=source_factory_id,
        target_controller=target_controller,
    )
    if target_factory_id is None:
        return {}, {"reason": "target_factory_not_found"}

    target_source = await afterbuy.get_products_by_factory_id(
        target_controller,
        int(target_factory_id),
    )
    target_items = [
        item.model_dump(mode="json", exclude_none=True)
        for item in target_source.products
    ]

    mapped: dict[str, str] = {}
    limit = min(len(products), len(target_items))
    for index in range(limit):
        product_ean = str(products[index].get("ean") or "").strip()
        source_ean = (
            _source_item_ean(source_items[index])
            if index < len(source_items)
            else ""
        )
        target_ean = _source_item_ean(target_items[index])
        current_ean = product_ean or source_ean
        if current_ean and target_ean:
            mapped[current_ean] = target_ean

    meta = {
        "source_factory_id": str(source_factory_id),
        "target_factory_id": target_factory_id,
        "target_account": target_controller.value,
        "source_items": len(source_items),
        "target_items": len(target_items),
        "products": len(products),
        "mapped": len(mapped),
    }
    MAPPER_LOGGER.info(
        "step=xl_ean_mapping_by_factory_index source_factory_id=%s target_factory_id=%s products=%s source_items=%s target_items=%s mapped=%s",
        source_factory_id,
        target_factory_id,
        len(products),
        len(source_items),
        len(target_items),
        len(mapped),
    )
    return mapped, meta


async def _build_factory_submit_preview_snapshots(
    *,
    process_id: str,
    payload: dict[str, Any],
    task: dict[str, Any] | None,
    afterbuy: AfterbuyService | None,
) -> dict[str, Any]:
    raw_products = payload.get("products")
    media_base_url = str(payload.get("media_base_url") or "").strip() or None
    if not isinstance(raw_products, list) or not raw_products:
        raise ValueError("products must be a non-empty array")

    products = [copy.deepcopy(item) for item in raw_products if isinstance(item, dict)]
    if not products:
        raise ValueError("products must contain valid product objects")

    controller_value = str(
        (task or {}).get("controller") or payload.get("controller") or "jv"
    ).lower()
    controller = Controller(controller_value)
    factory_id = str(
        (task or {}).get("factory_id") or payload.get("factory_id") or "unknown"
    )
    source_items_raw = (task or {}).get("source_items_raw")
    source_items = source_items_raw if isinstance(source_items_raw, list) else []
    products_with_variants = [item for item in products if active_variant_items(item)]
    ean_pool_assignments: list[dict[str, str]] = []

    try:
        ean_pool_assignments = await _fill_missing_variant_eans_from_pool(
            process_id=process_id,
            products=products_with_variants,
        )

        variant_validation_errors = validate_variant_export_identifiers(
            products_with_variants
        )
        if variant_validation_errors:
            return {
                "success": False,
                "variant_validation_errors": variant_validation_errors,
                "ean_pool_assignments": ean_pool_assignments,
            }

        expanded_products = expand_products_with_variants(
            products,
            media_base_url=media_base_url,
        )
        products_for_otto = [
            _shape_product_media_assets_for_otto(item)
            for item in await _translate_user_attribute_values_for_otto(
                expanded_products
            )
            if isinstance(item, dict)
        ]

        validated: list[dict[str, Any]] = []
        validated_models: list[ProductPayload] = []
        for item in products_for_otto:
            model = ProductPayload.model_validate(item)
            validated_models.append(model)
            validated_item = model.model_dump(mode="json", exclude_none=True)
            shipping_profile_id = str(
                item.get("shippingProfileID")
                or item.get("shippingProfileId")
                or item.get("shipping_profile_id")
                or ""
            ).strip()
            if shipping_profile_id:
                validated_item["shippingProfileID"] = shipping_profile_id
            validated.append(validated_item)

        xl_ean_map: dict[str, str] = {}
        xl_ean_map_meta: dict[str, Any] = {}
        xl_validated: list[dict[str, Any]] = []
        xl_validated_models: list[ProductPayload] = []
        if controller == Controller.JV and afterbuy is not None:
            xl_ean_map, xl_ean_map_meta = await _map_eans_from_counterpart_factory_by_index(
                afterbuy=afterbuy,
                source_factory_id=factory_id,
                target_controller=Controller.XL,
                source_items=source_items,
                products=validated,
            )
            xl_validated = _clone_products_with_account_map(
                validated,
                xl_ean_map,
                controller=Controller.XL,
            )
            xl_validated_models = [
                ProductPayload.model_validate(item) for item in xl_validated
            ]

        jv_snapshot_payloads = [
            _create_product_body(
                _apply_account_fields_to_product(item, controller=controller)
            )
            for item in validated
        ]
        jv_snapshot_path = _save_final_edited_payloads_snapshot(
            payloads=jv_snapshot_payloads,
            controller=controller,
            factory_id=factory_id,
            process_id=f"{process_id}_preview",
        )
        xl_snapshot_path = None
        if xl_validated:
            xl_snapshot_path = _save_final_edited_payloads_snapshot(
                payloads=[_create_product_body(item) for item in xl_validated],
                controller=Controller.XL,
                factory_id=factory_id,
                process_id=f"{process_id}_preview",
            )
        availability_payloads = [
            _availability_payload_item(controller=controller, product=item)
            for item in validated
        ] + [
            _availability_payload_item(controller=Controller.XL, product=item)
            for item in xl_validated
        ]
        availability_snapshot_path = _save_final_availability_snapshot(
            payloads=availability_payloads,
            factory_id=factory_id,
            process_id=f"{process_id}_preview",
        )

        return {
            "success": True,
            "process_id": process_id,
            "controller": controller.value,
            "factory_id": factory_id,
            "jv_snapshot_path": jv_snapshot_path.as_posix(),
            "xl_snapshot_path": xl_snapshot_path.as_posix()
            if xl_snapshot_path is not None
            else None,
            "availability_snapshot_path": availability_snapshot_path.as_posix(),
            "products_count": len(validated_models),
            "xl_products_count": len(xl_validated_models),
            "ean_pool_assignments": ean_pool_assignments,
            "ean_pool_preview_released": True,
            "xl_ean_map": xl_ean_map,
            "xl_ean_map_meta": xl_ean_map_meta,
            "sample": {
                "jv": [
                    {
                        "sku": item.get("sku"),
                        "ean": item.get("ean"),
                        "productReference": item.get("productReference"),
                    }
                    for item in validated[:5]
                ],
                "xl": [
                    {
                        "sku": item.get("sku"),
                        "ean": item.get("ean"),
                        "productReference": item.get("productReference"),
                    }
                    for item in xl_validated[:5]
                ],
            },
        }
    finally:
        await _release_ean_pool_assignments(ean_pool_assignments)


async def _run_factory_availability_task(
    *,
    process_id: str,
    availability_items: list[dict[str, Any]],
    product_service: ProductService,
) -> None:
    task = await _get_factory_task_state(process_id) or {
        "process_id": process_id,
        "status": "DONE",
    }
    MAPPER_LOGGER.info(
        "step=availability_background_delay_start process_id=%s delay_sec=%s products=%s",
        process_id,
        FACTORY_AVAILABILITY_AFTER_CREATE_DELAY_SEC,
        len(availability_items),
    )
    await asyncio.sleep(FACTORY_AVAILABILITY_AFTER_CREATE_DELAY_SEC)

    task["availability_state"] = "IN_PROGRESS"
    task["availability_started_at"] = datetime.now(UTC).isoformat()
    task["availability_progress_total"] = len(availability_items)
    task["availability_progress_completed"] = 0
    task["availability_progress_percent"] = 0
    task["updated_at"] = datetime.now(UTC).isoformat()
    task["heartbeat_at"] = task["updated_at"]
    await _save_factory_task_state(process_id, task)

    availability_errors: list[dict[str, str]] = []
    availability_queue: asyncio.Queue[tuple[int, dict[str, Any]] | None] = asyncio.Queue()
    availability_lock = asyncio.Lock()

    async def _submit_availability(index: int, item: dict[str, Any]) -> None:
        item_controller = Controller(str(item.get("controller") or "jv").lower())
        sku = str(item.get("sku") or "").strip()
        shipping_profile_id = str(item.get("shippingProfileID") or "").strip()
        quantity = str(item.get("quantity") or "20").strip() or "20"
        processing_time = str(item.get("processingTime") or "DEFAULT").strip() or "DEFAULT"
        MAPPER_LOGGER.info(
            "step=availability_background_item_start process_id=%s index=%s controller=%s sku=%s shipping_profile_id=%s quantity=%s",
            process_id,
            index,
            item_controller.value,
            sku or "-",
            shipping_profile_id or "-",
            quantity,
        )
        if not sku:
            raise ValueError("missing sku")
        if not shipping_profile_id:
            raise ValueError("missing shipping profile")

        availability_result = await product_service.create_availability(
            Availability(
                sku=sku,
                quantity=quantity,
                shippingProfileID=shipping_profile_id,
                processingTime=processing_time,
                controller=item_controller,
            )
        )
        quantity_ok = bool(
            availability_result.update_quantity
            and availability_result.update_quantity.success
        )
        delivery_ok = bool(
            availability_result.update_delivery
            and availability_result.update_delivery.success
        )
        if not quantity_ok or not delivery_ok:
            quantity_error = (
                availability_result.update_quantity.errors
                if availability_result.update_quantity
                else ""
            )
            delivery_error = (
                availability_result.update_delivery.errors
                if availability_result.update_delivery
                else ""
            )
            availability_errors.append(
                {
                    "variation": sku,
                    "code": "availability_failed",
                    "title": f"quantity={quantity_error or 'ok'}, delivery={delivery_error or 'ok'}",
                    "jsonPath": "availability",
                }
            )
            MAPPER_LOGGER.warning(
                "step=availability_background_item_result_failed process_id=%s index=%s controller=%s sku=%s quantity_ok=%s delivery_ok=%s quantity_error=%s delivery_error=%s",
                process_id,
                index,
                item_controller.value,
                sku,
                quantity_ok,
                delivery_ok,
                quantity_error or "",
                delivery_error or "",
            )

        async with availability_lock:
            current = await _get_factory_task_state(process_id) or task
            current["availability_progress_completed"] = (
                int(current.get("availability_progress_completed", 0)) + 1
            )
            current["availability_progress_percent"] = int(
                round(
                    (
                        current["availability_progress_completed"]
                        / max(1, len(availability_items))
                    )
                    * 100
                )
            )
            current["updated_at"] = datetime.now(UTC).isoformat()
            current["heartbeat_at"] = current["updated_at"]
            await _save_factory_task_state(process_id, current)

    async def _availability_worker(worker_id: int) -> None:
        while True:
            queued = await availability_queue.get()
            try:
                if queued is None:
                    return
                index, item = queued
                try:
                    await _submit_availability(index, item)
                except Exception as exc:
                    sku = str(item.get("sku") or "unknown").strip() or "unknown"
                    availability_errors.append(
                        {
                            "variation": sku,
                            "code": "availability_failed",
                            "title": str(exc),
                            "jsonPath": "availability",
                        }
                    )
                    MAPPER_LOGGER.exception(
                        "step=availability_background_item_failed process_id=%s worker=%s index=%s sku=%s error=%s",
                        process_id,
                        worker_id,
                        index,
                        sku,
                        exc,
                    )
                    async with availability_lock:
                        current = await _get_factory_task_state(process_id) or task
                        current["availability_progress_completed"] = (
                            int(current.get("availability_progress_completed", 0)) + 1
                        )
                        current["availability_progress_percent"] = int(
                            round(
                                (
                                    current["availability_progress_completed"]
                                    / max(1, len(availability_items))
                                )
                                * 100
                            )
                        )
                        current["updated_at"] = datetime.now(UTC).isoformat()
                        current["heartbeat_at"] = current["updated_at"]
                        await _save_factory_task_state(process_id, current)
            finally:
                availability_queue.task_done()

    for index, item in enumerate(availability_items):
        await availability_queue.put((index, item))

    availability_worker_count = min(
        FACTORY_PRODUCT_CONCURRENCY,
        len(availability_items),
    )
    availability_workers = [
        asyncio.create_task(_availability_worker(worker_id))
        for worker_id in range(availability_worker_count)
    ]
    for _ in availability_workers:
        await availability_queue.put(None)
    await availability_queue.join()
    await asyncio.gather(*availability_workers)

    current = await _get_factory_task_state(process_id) or task
    current["availability_errors_original"] = availability_errors
    current["availability_errors"] = (
        await _translate_otto_error_payload_for_ui(availability_errors)
        if availability_errors
        else availability_errors
    )
    current["availability_failed"] = len(availability_errors)
    current["availability_state"] = "FAILED" if availability_errors else "DONE"
    current["availability_finished_at"] = datetime.now(UTC).isoformat()
    current["availability_progress_completed"] = len(availability_items)
    current["availability_progress_total"] = len(availability_items)
    current["availability_progress_percent"] = 100
    current["updated_at"] = datetime.now(UTC).isoformat()
    current["heartbeat_at"] = current["updated_at"]
    await _save_factory_task_state(process_id, current)
    MAPPER_LOGGER.info(
        "step=availability_background_done process_id=%s products=%s failed=%s",
        process_id,
        len(availability_items),
        len(availability_errors),
    )


async def _run_factory_submit_task(
    *,
    process_id: str,
    payload: dict[str, Any],
    product_service: ProductService,
    afterbuy: AfterbuyService | None = None,
) -> None:
    task = await _get_factory_task_state(process_id)
    MAPPER_LOGGER.info(
        "step=submit_final_products_start process_id=%s has_task=%s",
        process_id,
        task is not None,
    )
    products = payload.get("products")
    media_base_url = str(payload.get("media_base_url") or "").strip() or None
    if not isinstance(products, list) or not products:
        MAPPER_LOGGER.error(
            "step=submit_final_products_failed process_id=%s error=empty_products",
            process_id,
        )
        return
    products = [copy.deepcopy(item) for item in products if isinstance(item, dict)]
    if not products:
        MAPPER_LOGGER.error(
            "step=submit_final_products_failed process_id=%s error=no_valid_products",
            process_id,
        )
        return

    validated: list[dict[str, Any]] = []
    validated_models: list[ProductPayload] = []
    xl_validated: list[dict[str, Any]] = []
    xl_validated_models: list[ProductPayload] = []
    xl_ean_map: dict[str, str] = {}
    xl_ean_map_meta: dict[str, Any] = {}
    ean_pool_assignments: list[dict[str, str]] = []
    controller_value = str(
        (task or {}).get("controller") or payload.get("controller") or "jv"
    ).lower()
    factory_id = str(
        (task or {}).get("factory_id") or payload.get("factory_id") or "unknown"
    )
    source_items_raw = (task or {}).get("source_items_raw")
    source_items = source_items_raw if isinstance(source_items_raw, list) else []

    try:
        controller = Controller(controller_value)
        if task is not None:
            task["status"] = "IN_PROGRESS"
            task["current_step"] = "final_validation_in_progress"
            task["updated_at"] = datetime.now(UTC).isoformat()
            task["heartbeat_at"] = task["updated_at"]
            await _save_factory_task_state(process_id, task)

        products_with_variants = [item for item in products if active_variant_items(item)]
        ean_pool_assignments = await _fill_missing_variant_eans_from_pool(
            process_id=process_id,
            products=products_with_variants,
        )
        if ean_pool_assignments:
            task["ean_pool_assignments"] = ean_pool_assignments
            await _save_factory_task_state(process_id, task)

        variant_validation_errors = validate_variant_export_identifiers(
            products_with_variants
        )
        if products_with_variants and not variant_validation_errors:
            expanded_for_conflicts = expand_products_with_variants(
                products_with_variants,
                media_base_url=media_base_url,
            )
            conflict_skus = {
                str(item.get("sku") or "").strip()
                for item in expanded_for_conflicts
                if str(item.get("sku") or "").strip()
            }
            conflict_eans = {
                str(item.get("ean") or "").strip()
                for item in expanded_for_conflicts
                if str(item.get("ean") or "").strip()
            }
            async with SessionLocal() as validation_session:
                try:
                    variant_validation_errors = await find_identifier_conflicts(
                        validation_session,
                        skus=conflict_skus,
                        eans=conflict_eans,
                    )
                except Exception as exc:
                    MAPPER_LOGGER.warning(
                        "step=variant_identifier_conflict_check_skipped process_id=%s error=%s",
                        process_id,
                        exc,
                    )
                    variant_validation_errors = []

        if variant_validation_errors:
            await _release_ean_pool_assignments(ean_pool_assignments)
            if task is None:
                task = {
                    "process_id": process_id,
                    "controller": controller.value,
                    "status": "FAILED",
                    "created_at": datetime.now(UTC).isoformat(),
                }
            task["status"] = "FAILED"
            task["current_step"] = "final_validation_failed"
            task["variant_validation_errors"] = variant_validation_errors
            task["otto_failed_result_original"] = {"results": []}
            task["otto_failed_result"] = {
                "results": [
                    {
                        "variation": item["variation"],
                        "errors": [
                            {
                                "code": item["code"],
                                "title": item["title"],
                                "jsonPath": item["jsonPath"],
                            }
                        ],
                    }
                    for item in variant_validation_errors
                ]
            }
            task["issues"] = [
                f"{item['variation']}: {item['code']}"
                for item in variant_validation_errors[:100]
            ]
            task["updated_at"] = datetime.now(UTC).isoformat()
            task["heartbeat_at"] = task["updated_at"]
            await _save_factory_task_state(process_id, task)
            return

        expanded_products = expand_products_with_variants(
            [item for item in products if isinstance(item, dict)],
            media_base_url=media_base_url,
        )

        products_for_otto = [
            _shape_product_media_assets_for_otto(item)
            for item in await _translate_user_attribute_values_for_otto(
                expanded_products
            )
            if isinstance(item, dict)
        ]

        for index, item in enumerate(products_for_otto):
            model = ProductPayload.model_validate(item)
            MAPPER_LOGGER.info(
                "step=submit_final_products_item_validated process_id=%s index=%s sku=%s category=%s",
                process_id,
                index,
                model.sku,
                model.productDescription.category,
            )
            validated_models.append(model)
            validated_item = model.model_dump(mode="json", exclude_none=True)
            shipping_profile_id = str(
                item.get("shippingProfileID")
                or item.get("shippingProfileId")
                or item.get("shipping_profile_id")
                or ""
            ).strip()
            if shipping_profile_id:
                validated_item["shippingProfileID"] = shipping_profile_id
            else:
                MAPPER_LOGGER.warning(
                    "step=submit_final_products_missing_shipping_profile process_id=%s index=%s sku=%s source_keys=%s",
                    process_id,
                    index,
                    model.sku,
                    sorted(str(key) for key in item.keys()),
                )
            validated.append(validated_item)

        if controller == Controller.JV and afterbuy is not None:
            try:
                xl_ean_map, xl_ean_map_meta = await _map_eans_from_counterpart_factory_by_index(
                    afterbuy=afterbuy,
                    source_factory_id=factory_id,
                    target_controller=Controller.XL,
                    source_items=source_items,
                    products=validated,
                )
            except Exception as exc:
                MAPPER_LOGGER.warning(
                    "step=xl_ean_mapping_failed process_id=%s error=%s",
                    process_id,
                    exc,
                )
                xl_ean_map = {}
                xl_ean_map_meta = {"error": str(exc)}
            xl_validated = _clone_products_with_account_map(
                validated,
                xl_ean_map,
                controller=Controller.XL,
            )
            for index, item in enumerate(xl_validated):
                model = ProductPayload.model_validate(item)
                xl_validated_models.append(model)
                MAPPER_LOGGER.info(
                    "step=submit_final_products_xl_item_validated process_id=%s index=%s sku=%s source_ean=%s xl_ean=%s",
                    process_id,
                    index,
                    model.sku,
                    validated[index].get("ean") if index < len(validated) else "",
                    item.get("ean"),
                )

        jv_snapshot_payloads = [
            _create_product_body(
                _apply_account_fields_to_product(item, controller=controller)
            )
            for item in validated
        ]
        file_path = _save_final_edited_payloads_snapshot(
            payloads=jv_snapshot_payloads,
            controller=controller,
            factory_id=factory_id,
            process_id=process_id,
        )
        xl_file_path = None
        if xl_validated:
            xl_file_path = _save_final_edited_payloads_snapshot(
                payloads=[_create_product_body(item) for item in xl_validated],
                controller=Controller.XL,
                factory_id=factory_id,
                process_id=process_id,
            )
        availability_payloads = [
            _availability_payload_item(controller=controller, product=item)
            for item in validated
        ] + [
            _availability_payload_item(controller=Controller.XL, product=item)
            for item in xl_validated
        ]
        availability_file_path = _save_final_availability_snapshot(
            payloads=availability_payloads,
            factory_id=factory_id,
            process_id=process_id,
        )
        if task is None:
            task = {
                "process_id": process_id,
                "controller": controller.value,
                "factory_id": factory_id,
                "status": "IN_PROGRESS",
                "created_at": datetime.now(UTC).isoformat(),
            }
        task["final_snapshot_path"] = file_path.as_posix()
        if xl_file_path is not None:
            task["final_snapshot_path_xl"] = xl_file_path.as_posix()
        task["availability_snapshot_path"] = availability_file_path.as_posix()
        task["xl_ean_map"] = xl_ean_map
        task["xl_ean_map_meta"] = xl_ean_map_meta
        task["xl_products_count"] = len(xl_validated_models)
        task["final_products_count"] = len(validated)
        task["current_step"] = "otto_create_in_progress"
        task["updated_at"] = datetime.now(UTC).isoformat()
        task["heartbeat_at"] = task["updated_at"]
        await _save_factory_task_state(process_id, task)

        MAPPER_LOGGER.info(
            "step=submit_final_products_create_start process_id=%s products=%s",
            process_id,
            len(validated_models),
        )
        (
            otto_process_id,
            otto_state,
            update_result,
            failed_result,
        ) = await _submit_products_to_otto_in_batches(
            product_service=product_service,
            controller=controller,
            products=validated_models,
        )
        xl_otto_process_id = None
        xl_otto_state = None
        xl_update_result = None
        xl_failed_result = None
        if xl_validated_models:
            (
                xl_otto_process_id,
                xl_otto_state,
                xl_update_result,
                xl_failed_result,
            ) = await _submit_products_to_otto_in_batches(
                product_service=product_service,
                controller=Controller.XL,
                products=xl_validated_models,
            )

        task["otto_process_id"] = otto_process_id
        task["otto_create_state"] = otto_state
        task["otto_update_result"] = update_result
        task["otto_failed_result_original"] = failed_result
        task["otto_failed_result"] = (
            await _translate_otto_error_payload_for_ui(failed_result)
            if failed_result
            else failed_result
        )
        failed_issues = _extract_otto_error_issues(
            task["otto_failed_result"] or failed_result
        )
        if failed_issues:
            task["issues"] = failed_issues[:100]
        if xl_validated_models:
            task["otto_xl_process_id"] = xl_otto_process_id
            task["otto_xl_create_state"] = xl_otto_state
            task["otto_xl_update_result"] = xl_update_result
            task["otto_xl_failed_result_original"] = xl_failed_result
            task["otto_xl_failed_result"] = (
                await _translate_otto_error_payload_for_ui(xl_failed_result)
                if xl_failed_result
                else xl_failed_result
            )
            xl_failed_issues = _extract_otto_error_issues(
                task["otto_xl_failed_result"] or xl_failed_result
            )
            if xl_failed_issues:
                task["issues"] = (task.get("issues") or []) + xl_failed_issues[:100]
        task["updated_at"] = datetime.now(UTC).isoformat()
        task["heartbeat_at"] = task["updated_at"]
        create_failed_count = int((update_result or {}).get("failed") or 0)
        xl_create_failed_count = int((xl_update_result or {}).get("failed") or 0)
        task["status"] = (
            "DONE"
            if create_failed_count == 0 and xl_create_failed_count == 0
            else "FAILED"
        )
        task["current_step"] = "otto_create_done"
        task["products_count"] = len(validated_models) + len(xl_validated_models)
        jv_create_done = (
            create_failed_count == 0
            and int((update_result or {}).get("succeeded") or 0) >= len(validated_models)
        )
        if jv_create_done:
            await _mark_ean_pool_assignments_used(ean_pool_assignments)
        else:
            await _release_ean_pool_assignments(ean_pool_assignments)
        await _save_factory_task_state(process_id, task)

        succeeded_count = int((update_result or {}).get("succeeded") or 0)
        xl_succeeded_count = int((xl_update_result or {}).get("succeeded") or 0)
        xl_create_done = not xl_validated_models or xl_succeeded_count >= len(
            xl_validated_models
        )
        if (
            task["status"] != "DONE"
            or succeeded_count < len(validated_models)
            or not xl_create_done
        ):
            MAPPER_LOGGER.warning(
                "step=submit_availability_skipped process_id=%s status=%s succeeded=%s expected=%s xl_succeeded=%s xl_expected=%s update_result=%s xl_update_result=%s",
                process_id,
                task["status"],
                succeeded_count,
                len(validated_models),
                xl_succeeded_count,
                len(xl_validated_models),
                update_result,
                xl_update_result,
            )
        elif (
            succeeded_count >= len(validated_models)
            and xl_create_done
        ):
            task["availability_state"] = "QUEUED"
            task["availability_progress_total"] = len(availability_payloads)
            task["availability_progress_completed"] = 0
            task["availability_progress_percent"] = 0
            task["availability_queued_at"] = datetime.now(UTC).isoformat()
            await _save_factory_task_state(process_id, task)
            MAPPER_LOGGER.info(
                "step=availability_background_enqueue process_id=%s delay_sec=%s products=%s",
                process_id,
                FACTORY_AVAILABILITY_AFTER_CREATE_DELAY_SEC,
                len(availability_payloads),
            )
            await enqueue_job(
                "submit_factory_availability_task",
                process_id=process_id,
                availability_items=availability_payloads,
            )

        MAPPER_LOGGER.info(
            "step=submit_final_products_done process_id=%s status=%s products=%s",
            process_id,
            task["status"],
            len(validated_models),
        )
    except Exception as exc:
        await _release_ean_pool_assignments(ean_pool_assignments)
        current = task or {
            "process_id": process_id,
            "status": "FAILED",
        }
        current["status"] = "FAILED"
        current["current_step"] = "otto_create_failed"
        current["updated_at"] = datetime.now(UTC).isoformat()
        current["heartbeat_at"] = current["updated_at"]
        current["issues"] = [str(exc)]
        await _save_factory_task_state(process_id, current)
        MAPPER_LOGGER.exception(
            "step=submit_final_products_failed process_id=%s error=%s",
            process_id,
            exc,
        )


@router.post("/tasks/create-from-factory/{process_id}/submit-preview")
async def preview_factory_submit_payloads(
    process_id: str,
    request: Request,
    payload: dict[str, Any] = Body(default_factory=dict),
    afterbuy: AfterbuyService = Depends(get_afterbuy_login),
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    if not settings.factory_submit_preview_enabled:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Submit preview endpoint is disabled.",
        )

    task = await _get_owned_factory_task_state(process_id, current_user.id)
    if task is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Task not found.",
        )

    preview_payload = dict(payload or {})
    if not isinstance(preview_payload.get("products"), list):
        preview_payload["products"] = task.get("products") or task.get("submitted_products")
    preview_payload["controller"] = (
        preview_payload.get("controller") or task.get("controller") or "jv"
    )
    preview_payload["factory_id"] = (
        preview_payload.get("factory_id") or task.get("factory_id") or "unknown"
    )
    preview_payload["media_base_url"] = (
        str(preview_payload.get("media_base_url") or "").strip()
        or str(request.base_url).rstrip("/")
    )

    try:
        result = await _build_factory_submit_preview_snapshots(
            process_id=process_id,
            payload=preview_payload,
            task=task,
            afterbuy=afterbuy,
        )
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(exc),
        ) from exc
    except Exception as exc:
        MAPPER_LOGGER.exception(
            "step=submit_preview_failed process_id=%s error=%s",
            process_id,
            exc,
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(exc),
        ) from exc

    now = datetime.now(UTC).isoformat()
    task["submit_preview"] = {
        "created_at": now,
        "jv_snapshot_path": result.get("jv_snapshot_path"),
        "xl_snapshot_path": result.get("xl_snapshot_path"),
        "products_count": result.get("products_count"),
        "xl_products_count": result.get("xl_products_count"),
        "xl_ean_map": result.get("xl_ean_map") or {},
        "xl_ean_map_meta": result.get("xl_ean_map_meta") or {},
        "ean_pool_assignments": result.get("ean_pool_assignments") or [],
        "ean_pool_preview_released": result.get("ean_pool_preview_released"),
    }
    task["updated_at"] = now
    await _save_factory_task_state(process_id, task)
    return result


@router.post("/tasks/create-from-factory/{process_id}/submit")
async def submit_factory_prepared_products(
    process_id: str,
    payload: dict[str, Any],
    request: Request,
    current_user=Depends(require_role([RoleEnum.SEO])),
):
    task = await _get_owned_factory_task_state(process_id, current_user.id)
    MAPPER_LOGGER.info(
        "step=submit_final_products_queue_start process_id=%s has_task=%s",
        process_id,
        task is not None,
    )
    if task is None:
        return JSONResponse(
            status_code=status.HTTP_404_NOT_FOUND,
            content={
                "success": False,
                "message": "Task not found",
                "process_id": process_id,
            },
        )
    products = payload.get("products")
    if not isinstance(products, list) or not products:
        return JSONResponse(
            status_code=status.HTTP_400_BAD_REQUEST,
            content={"success": False, "message": "products must be a non-empty array"},
        )

    controller_value = str(
        (task or {}).get("controller") or payload.get("controller") or "jv"
    ).lower()
    try:
        controller = Controller(controller_value)
    except ValueError:
        return JSONResponse(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            content={
                "success": False,
                "message": f"Invalid controller: {controller_value}",
                "process_id": process_id,
            },
        )

    if task is not None:
        submitted_products = [item for item in products if isinstance(item, dict)]
        media_base_url = (
            str(payload.get("media_base_url") or "").strip()
            or str(request.base_url).rstrip("/")
        )
        queued_products_count = len(
            expand_products_with_variants(
                submitted_products,
                media_base_url=media_base_url,
            )
        )
        now = datetime.now(UTC).isoformat()
        task["status"] = "IN_PROGRESS"
        task["current_step"] = "otto_create_queued"
        task["current_step_started_at"] = now
        task["updated_at"] = now
        task["heartbeat_at"] = now
        task["progress_total"] = queued_products_count
        task["progress_completed"] = 0
        task["progress_percent"] = 0
        task["products"] = submitted_products
        task["submitted_products"] = submitted_products
        await _save_factory_task_state(
            process_id,
            task,
            created_by_user_id=current_user.id,
        )

    payload_for_job = dict(payload)
    payload_for_job["media_base_url"] = (
        str(payload.get("media_base_url") or "").strip()
        or str(request.base_url).rstrip("/")
    )

    await enqueue_job(
        "submit_factory_products_task",
        process_id=process_id,
        payload=payload_for_job,
    )

    return {
        "success": True,
        "process_id": process_id,
        "process_state": "IN_PROGRESS",
        "queued": True,
        "products_count": len(
            expand_products_with_variants(
                [item for item in products if isinstance(item, dict)],
                media_base_url=payload_for_job["media_base_url"],
            )
        ),
    }


@router.get("/fetch-otto-categories-to-db")
async def fetch_otto_categories_to_db(
    product_service: ProductService = Depends(get_product_service),
    session: AsyncSession = Depends(get_db),
):
    await product_service.fetch_all_categories_to_db(session)

    return {"success": True, "message": "Category sync started"}
